# -*- coding: utf-8 -*-
"""
UXM Stage Runner
================
Windows uzerinde UXM/uXBasic stage test hattini tek komutla calistirir:
1) Stage numarasini otomatik bulur.
2) En guncel smoke test bat dosyasini bulup once calistirir.
3) Derleyiciyi derler.
4) uxm/tests altindaki tum .uxm testlerini sirayla derler, calistirir ve olcer.
5) test_history.csv ve test_stats_summary.csv dosyalarini gunceller.
6) ASM optimizer hattini calistirir: build/asm -> yeni_optimize_asm -> *_opt.exe -> orijinal/opt cikti ve sure karsilastirma.
7) Son aktif build klasorunu "build stage N" adiyla arsivler.

Kullanim:
    py -3 UXM_STAGE_RUNNER.py --stage auto
    py -3 UXM_STAGE_RUNNER.py --stage 12
    py -3 UXM_STAGE_RUNNER.py --stage auto --skip-opt

Sadece standart Python kutuphanelerini kullanir. Excel raporu icin openpyxl varsa xlsx de yazar, yoksa CSV/MD ile devam eder.
"""
from __future__ import annotations

import argparse
import csv
import datetime as _dt
import difflib
import json
import math
import os
import re
import shutil
import sqlite3
import statistics
import subprocess
import sys
import time
from collections import Counter, defaultdict
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Iterable, Optional


# -----------------------------
# Genel yardimcilar
# -----------------------------

STAGE_STATE_FILE = "stage_state.json"
DEFAULT_TEST_DIRS = [
    Path("uxm/tests/fp"),
    Path("uxm/tests/math"),
    Path("uxm/tests/matrix"),
    Path("uxm/tests/native"),
    Path("uxm/tests/v33"),
]


@dataclass
class CmdResult:
    label: str
    command: str
    returncode: int
    elapsed_sec: float
    stdout: str
    stderr: str


@dataclass
class TestResult:
    stage: int
    timestamp: str
    test_path: str
    test_name: str
    status: str
    returncode: int
    duration_sec: float
    build_sec: float
    output_compact: str
    log_file: str


@dataclass
class OptResult:
    stage: int
    timestamp: str
    test_name: str
    status: str
    orig_ms: Optional[float]
    opt_ms: Optional[float]
    gain_percent: Optional[float]
    output_match: Optional[bool]
    returncode_orig: Optional[int]
    returncode_opt: Optional[int]
    diff_file: str


class StageRunner:
    def __init__(self, args: argparse.Namespace):
        self.args = args
        self.base = Path.cwd()
        self.now = _dt.datetime.now()
        self.run_stamp = self.now.strftime("%Y%m%d_%H%M%S")
        self.stage = self.detect_stage(args.stage)
        self.result_log = self.next_result_log()
        self.run_dir = self.base / "stage_runs" / f"stage_{self.stage}_{self.run_stamp}"
        self.log_dir = self.run_dir / "logs"
        self.per_test_log_dir = self.log_dir / "tests"
        self.opt_log_dir = self.log_dir / "optimizer"
        self.report_dir = self.base / "optimizasyon"
        for d in [self.run_dir, self.log_dir, self.per_test_log_dir, self.opt_log_dir, self.report_dir]:
            d.mkdir(parents=True, exist_ok=True)
        self.fbc = self.detect_fbc(args.fbc)
        self.nasm = args.nasm or os.environ.get("UXM_NASM") or "nasm"
        self.session_start = time.perf_counter()
        self.build_duration = 0.0
        self.test_results: list[TestResult] = []
        self.opt_results: list[OptResult] = []
        self.summary_lines: list[str] = []

    # -----------------------------
    # Tespitler
    # -----------------------------
    def detect_stage(self, requested: str) -> int:
        if requested and requested.lower() != "auto":
            return int(requested)

        candidates: list[int] = []

        state_path = self.base / STAGE_STATE_FILE
        if state_path.exists():
            try:
                data = json.loads(state_path.read_text(encoding="utf-8"))
                if isinstance(data.get("last_completed_stage"), int):
                    candidates.append(data["last_completed_stage"] + 1)
                if isinstance(data.get("current_stage"), int):
                    candidates.append(data["current_stage"])
            except Exception:
                pass

        # "build stage 11", "build_stage11", "build-stage-11", "built stage 11" gibi klasorler.
        build_re = re.compile(r"^(?:build|built)[ _-]*(?:stage)?[ _-]*(\d+)$", re.IGNORECASE)
        for p in self.base.iterdir():
            if p.is_dir():
                m = build_re.match(p.name.strip())
                if m:
                    candidates.append(int(m.group(1)) + 1)

        # Smoke dosyalari genellikle son tamamlanan stage'i gosterir: run_stage11_smoke.bat -> siradaki 12.
        smoke_re = re.compile(r"run_stage(\d+)_smoke\.bat$", re.IGNORECASE)
        for p in self.base.glob("run_stage*_smoke.bat"):
            m = smoke_re.match(p.name)
            if m:
                candidates.append(int(m.group(1)) + 1)

        # Hicbir sey yoksa Stage 1.
        return max(candidates) if candidates else 1

    def next_result_log(self) -> Path:
        nums = []
        for p in self.base.glob("sonuc*.txt"):
            m = re.fullmatch(r"sonuc(\d+)\.txt", p.name, re.IGNORECASE)
            if m:
                nums.append(int(m.group(1)))
        n = (max(nums) + 1) if nums else 1
        return self.base / f"sonuc{n}.txt"

    def detect_fbc(self, explicit: Optional[str]) -> str:
        candidates = []
        if explicit:
            candidates.append(explicit)
        for env_name in ("UXM_FBC", "FBC"):
            val = os.environ.get(env_name)
            if val:
                candidates.append(val)
        # Kullanici makinesindeki bilinen eski yol. Varsa kullanilir; yoksa PATH'teki fbc'ye duser.
        candidates += [
            r"C:\Users\mete\Downloads\BasicOyunSource\uXBasic_repo\tools\FreeBASIC-1.10.1-win64\fbc.exe",
            r"tools\FreeBASIC-1.10.1-win64\fbc.exe",
            r"FreeBASIC-1.10.1-win64\fbc.exe",
            "fbc",
        ]
        for c in candidates:
            if c.lower() == "fbc" or Path(c).exists():
                return c
        return "fbc"

    def discover_smoke_bats(self) -> list[Path]:
        bats = sorted(self.base.glob("*smoke*.bat"), key=lambda p: p.name.lower())

        def stage_no(p: Path) -> int:
            m = re.search(r"stage(\d+)", p.name, re.IGNORECASE)
            return int(m.group(1)) if m else -1

        if self.args.all_smoke:
            return sorted(bats, key=lambda p: (stage_no(p), p.name.lower()))
        return [max(bats, key=lambda p: (stage_no(p), p.stat().st_mtime))] if bats else []

    def discover_tests(self) -> list[Path]:
        tests: list[Path] = []
        roots = []
        for d in DEFAULT_TEST_DIRS:
            if (self.base / d).exists():
                roots.append(self.base / d)
        # Yeni klasor eklenirse de yakalansin.
        generic_root = self.base / "uxm" / "tests"
        if generic_root.exists():
            for child in sorted(generic_root.iterdir()):
                if child.is_dir() and child not in roots:
                    roots.append(child)

        for root in roots:
            for p in sorted(root.rglob("*.uxm")):
                if not self.args.include_tmp and p.name.lower().startswith("_tmp"):
                    continue
                tests.append(p)
        return tests

    # -----------------------------
    # Calistirma / log
    # -----------------------------
    def append_session_log(self, text: str) -> None:
        with self.result_log.open("a", encoding="utf-8", newline="") as f:
            f.write(text)
            if not text.endswith("\n"):
                f.write("\n")

    def run_cmd(self, cmd: str, label: str, cwd: Optional[Path] = None) -> CmdResult:
        start = time.perf_counter()
        proc = subprocess.run(
            cmd,
            cwd=str(cwd or self.base),
            shell=True,
            capture_output=True,
            text=True,
            errors="replace",
        )
        elapsed = time.perf_counter() - start
        return CmdResult(label, cmd, proc.returncode, elapsed, proc.stdout or "", proc.stderr or "")

    @staticmethod
    def compact_output(stdout: str, limit: int = 240) -> str:
        # Derleyici bilgi satirlarini ayiklayip test ciktisini kisa tutar.
        lines = []
        skip_prefix = (
            "ASM uretildi:",
            "[V3.3",
            "NASM:",
            "FreeBASIC runtime",
            "C:\\",
            "nasm ",
            "OK:",
        )
        for raw in stdout.splitlines():
            s = raw.strip()
            if not s:
                continue
            if s.startswith(skip_prefix):
                continue
            if s == "[UXM program derlendi.]":
                continue
            lines.append(s)
        joined = " | ".join(lines)
        return joined[:limit]

    def write_cmd_log(self, path: Path, result: CmdResult) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", encoding="utf-8", newline="") as f:
            f.write(f"LABEL: {result.label}\n")
            f.write(f"COMMAND: {result.command}\n")
            f.write(f"RETURNCODE: {result.returncode}\n")
            f.write(f"ELAPSED_SEC: {result.elapsed_sec:.6f}\n")
            f.write("\n--- STDOUT ---\n")
            f.write(result.stdout)
            f.write("\n--- STDERR ---\n")
            f.write(result.stderr)

    # -----------------------------
    # Fazlar
    # -----------------------------
    def init_log(self) -> None:
        self.result_log.write_text("", encoding="utf-8")
        self.append_session_log(f"START_SESSION@{self.now.strftime('%Y-%m-%d')}@{self.now.strftime('%H:%M:%S')}@STAGE@{self.stage}")
        self.append_session_log(f"RUN_DIR@{self.run_dir}")
        self.append_session_log(f"FBC@{self.fbc}")
        self.append_session_log(f"NASM@{self.nasm}")

    def build_compiler(self) -> bool:
        self.append_session_log(f"START_BUILD@{_dt.datetime.now().strftime('%H:%M:%S')}")
        result = self.run_cmd("call build_native.bat", "build_native")
        self.build_duration = result.elapsed_sec
        self.write_cmd_log(self.log_dir / "build_native.log", result)
        self.append_session_log(result.stdout)
        if result.stderr:
            self.append_session_log("--- STDERR ---")
            self.append_session_log(result.stderr)
        self.append_session_log(f"END_BUILD@{_dt.datetime.now().strftime('%H:%M:%S')}@SECONDS@{self.build_duration:.6f}")
        if result.returncode != 0:
            self.summary_lines.append(f"[FAIL] build_native.bat returncode={result.returncode}")
            return False
        self.summary_lines.append(f"[OK] compiler build: {self.build_duration:.3f} sn")
        return True

    def run_smoke_phase(self) -> bool:
        bats = self.discover_smoke_bats()
        if not bats:
            self.summary_lines.append("[WARN] Smoke bat bulunamadi; otomatik mini smoke uygulanacak.")
            return self.run_derived_smoke()

        ok = True
        for bat in bats:
            label = f"smoke:{bat.name}"
            self.append_session_log(f"START_SMOKE@{bat.name}@{_dt.datetime.now().strftime('%H:%M:%S')}")
            result = self.run_cmd(f"call {quote_win(bat.name)}", label)
            self.write_cmd_log(self.log_dir / f"{safe_stem(bat.stem)}.log", result)
            self.append_session_log(result.stdout)
            if result.stderr:
                self.append_session_log("--- SMOKE STDERR ---")
                self.append_session_log(result.stderr)
            self.append_session_log(f"END_SMOKE@{bat.name}@{_dt.datetime.now().strftime('%H:%M:%S')}@RC@{result.returncode}@SECONDS@{result.elapsed_sec:.6f}")
            if result.returncode != 0:
                ok = False
                self.summary_lines.append(f"[FAIL] smoke: {bat.name} ({result.elapsed_sec:.3f} sn)")
                if not self.args.continue_on_fail:
                    return False
            else:
                self.summary_lines.append(f"[OK] smoke: {bat.name} ({result.elapsed_sec:.3f} sn)")
        return ok

    def run_derived_smoke(self) -> bool:
        # Smoke dosyasi yoksa kararlı ve temsilci testlerden mini liste.
        expected = [
            (Path("uxm/tests/fp/test_fp01_add_int.uxm"), "46"),
            (Path("uxm/tests/native/test05_meta_add.uxm"), "30"),
            (Path("uxm/tests/v33/test_v33_tensor3d_index_slice.uxm"), "772377"),
        ]
        ok = True
        for rel, needle in expected:
            src = self.base / rel
            if not src.exists():
                self.summary_lines.append(f"[WARN] derived smoke atlandi, yok: {rel}")
                continue
            result = self.run_cmd(f"call build_one_native.bat {quote_win(str(rel))} -x", f"derived_smoke:{rel.name}")
            self.write_cmd_log(self.log_dir / f"derived_smoke_{safe_stem(rel.stem)}.log", result)
            if result.returncode != 0 or needle not in result.stdout:
                ok = False
                self.summary_lines.append(f"[FAIL] derived smoke: {rel} expected={needle}")
                if not self.args.continue_on_fail:
                    return False
            else:
                self.summary_lines.append(f"[OK] derived smoke: {rel} expected={needle}")
        return ok

    def run_full_tests(self) -> bool:
        tests = self.discover_tests()
        if not tests:
            self.summary_lines.append("[FAIL] .uxm test bulunamadi.")
            return False

        ok = True
        self.summary_lines.append(f"[INFO] tam test adedi: {len(tests)}")
        for idx, src_abs in enumerate(tests, 1):
            rel = src_abs.relative_to(self.base)
            rel_s = str(rel).replace("/", "\\")
            self.append_session_log("---------------------------------------")
            self.append_session_log(".T..E..S..T....B..A..S..I..............")
            self.append_session_log(f"DATA_START@{rel_s}@{_dt.datetime.now().strftime('%H:%M:%S')}")

            # Tam testte -x kullanmiyoruz: her test kendi adiyla build/asm,obj,exe altina kalsin.
            cmd = f"call build_one_native.bat {quote_win(rel_s)}"
            result = self.run_cmd(cmd, f"test:{rel_s}")
            test_log = self.per_test_log_dir / f"{idx:03d}_{safe_stem(src_abs.stem)}.log"
            self.write_cmd_log(test_log, result)

            self.append_session_log(result.stdout)
            if result.stderr:
                self.append_session_log("--- STDERR ---")
                self.append_session_log(result.stderr)

            status = "BASARILI" if result.returncode == 0 else "BASARISIZ"
            if result.returncode != 0:
                ok = False
            compact = self.compact_output(result.stdout)
            tr = TestResult(
                stage=self.stage,
                timestamp=_dt.datetime.now().isoformat(timespec="seconds"),
                test_path=rel_s,
                test_name=src_abs.stem,
                status=status,
                returncode=result.returncode,
                duration_sec=round(result.elapsed_sec, 6),
                build_sec=round(self.build_duration, 6),
                output_compact=compact,
                log_file=str(test_log.relative_to(self.base)).replace("/", "\\"),
            )
            self.test_results.append(tr)

            self.append_session_log(f"RESULT@[{status}]@{rel_s}")
            self.append_session_log(f"DATA_END@{rel_s}@{_dt.datetime.now().strftime('%H:%M:%S')}@SECONDS@{result.elapsed_sec:.6f}")
            self.append_session_log(".......................................")
            self.append_session_log("-T--E--S--T----S--O--N--U--------------")
            print(f"[{idx:03d}/{len(tests)}] {status:8s} {rel_s} ({result.elapsed_sec:.2f} sn)")

            if result.returncode != 0 and not self.args.continue_on_fail:
                break

        return ok

    # -----------------------------
    # ASM optimizer fazi
    # -----------------------------
    def optimizer_phase(self) -> bool:
        if self.args.skip_opt:
            self.summary_lines.append("[SKIP] optimizer fazi atlandi (--skip-opt).")
            return True
        build_asm = self.base / "build" / "asm"
        if not build_asm.exists():
            self.summary_lines.append("[WARN] build/asm yok; optimizer fazi atlandi.")
            return False

        self.summary_lines.append("[INFO] optimizer fazi basladi.")
        self.asm_intelligence_pass(build_asm)
        self.heavy_suggestion_pass()
        self.build_optimized_asm()
        self.compare_optimized_outputs()
        self.write_optimization_reports()
        return True

    def asm_intelligence_pass(self, asm_dir: Path) -> None:
        opt_dir = self.base / "yeni_optimize_asm"
        opt_dir.mkdir(exist_ok=True)
        report = self.report_dir / "asm_intel_report.txt"
        asm_files = sorted(asm_dir.glob("*.asm"))
        with report.open("w", encoding="utf-8", newline="") as rep:
            rep.write(f"UXM ASM INTELLIGENCE REPORT - Stage {self.stage}\n")
            rep.write(f"Kaynak: {asm_dir}\n")
            rep.write(f"ASM dosya sayisi: {len(asm_files)}\n")
            rep.write("=" * 80 + "\n")
            for asm in asm_files:
                text = asm.read_text(encoding="utf-8", errors="replace").splitlines()
                clean = [clean_instr(x) for x in text]
                clean = [x for x in clean if x]
                bigrams = [f"{clean[i]} -> {clean[i+1]}" for i in range(max(0, len(clean)-1))]
                jump_density = len([x for x in clean if x.lower().startswith(("j", "call"))]) / (len(clean) + 1)
                rep.write(f"\nANALIZ: {asm.name}\n")
                rep.write("-" * 60 + "\n")
                rep.write("Sik Kullanilan Komutlar:\n")
                for instr, cnt in Counter(clean).most_common(8):
                    rep.write(f" - {cnt:4d} kez: {instr}\n")
                rep.write("Sik Kullanilan Bigramlar:\n")
                for bg, cnt in Counter(bigrams).most_common(8):
                    rep.write(f" - {cnt:4d} kez: {bg}\n")
                rep.write(f"Jump/Call Yogunlugu: %{jump_density*100:.2f}\n")

                optimized = apply_safe_asm_rules(text)
                (opt_dir / asm.name).write_text("\n".join(optimized) + "\n", encoding="utf-8")
        self.summary_lines.append(f"[OK] ASM intelligence: {report}")

    def heavy_suggestion_pass(self) -> None:
        intel_report = self.report_dir / "asm_intel_report.txt"
        out1 = self.report_dir / "sihirli_asm_onerileri.txt"
        out2 = self.report_dir / "strateji_kitabi_v2.txt"
        intel = intel_report.read_text(encoding="utf-8", errors="replace") if intel_report.exists() else ""
        suggestions = [
            f"=== UXM STAGE-{self.stage} ASM OPTIMIZASYON STRATEJI KITABI ===",
            "Durum: Orijinal build/asm dosyalarina dokunulmaz; optimize kopyalar yeni_optimize_asm altina yazilir.",
            "",
        ]
        if "push rax" in intel and "pop rax" in intel:
            suggestions += [
                "[KRITIK] Gereksiz push/pop ve register golgeleme adaylari var.",
                "Cozum: ABI'yi bozmadan r8-r11 gibi volatile registerlari gecici depo olarak kullan; stack trafiğini azalt.",
                "",
            ]
        if "mov rax, 0" in intel:
            suggestions += ["[BASIT] mov rax, 0 -> xor rax, rax adaylari var.", ""]
        if "cmp" in intel and "jae" in intel:
            suggestions += [
                "[DONGU] Sinir kontrolu / branch yogunlugu goruluyor.",
                "Cozum: Tensor/matrix dongulerinde toplam boyutu dongu oncesi dogrula, ic dongudeki kontrolu azalt.",
                "",
            ]
        suggestions += [
            "=== SEARCH/REPLACE ADAYLARI ===",
            "mov rax, 0  -> xor rax, rax",
            "add reg, 1  -> inc reg    ; bayrak etkisi kabul ediliyorsa",
            "sub reg, 1  -> dec reg    ; bayrak etkisi kabul ediliyorsa",
            "imul reg, reg, 2/4/8/16 -> lea/shl ; sadece semantik esitse",
        ]
        out1.write_text("\n".join(suggestions) + "\n", encoding="utf-8")
        out2.write_text("\n".join(suggestions) + "\n", encoding="utf-8")
        self.summary_lines.append(f"[OK] heavy suggestions: {out1.name}, {out2.name}")

    def build_optimized_asm(self) -> None:
        opt_dir = self.base / "yeni_optimize_asm"
        obj_dir = self.base / "build" / "obj"
        exe_dir = self.base / "build" / "exe"
        runtime = self.base / "uxm" / "core" / "runtime" / "uxm31_runtime_fb_full.bas"
        obj_dir.mkdir(parents=True, exist_ok=True)
        exe_dir.mkdir(parents=True, exist_ok=True)
        for asm in sorted(opt_dir.glob("*.asm")):
            name = asm.stem
            obj = obj_dir / f"{name}_opt.o"
            exe = exe_dir / f"{name}_opt.exe"
            nasm_cmd = f"{quote_win(self.nasm)} -f win64 {quote_win(str(asm))} -o {quote_win(str(obj))}"
            r1 = self.run_cmd(nasm_cmd, f"nasm_opt:{name}")
            self.write_cmd_log(self.opt_log_dir / f"{safe_stem(name)}_nasm.log", r1)
            if r1.returncode != 0:
                self.opt_results.append(OptResult(self.stage, _dt.datetime.now().isoformat(timespec="seconds"), name, "NASM_FAIL", None, None, None, None, None, None, ""))
                continue
            link_cmd = f"{quote_win(self.fbc)} {quote_win(str(runtime))} {quote_win(str(obj))} -x {quote_win(str(exe))}"
            r2 = self.run_cmd(link_cmd, f"link_opt:{name}")
            self.write_cmd_log(self.opt_log_dir / f"{safe_stem(name)}_link.log", r2)
            if r2.returncode != 0:
                self.opt_results.append(OptResult(self.stage, _dt.datetime.now().isoformat(timespec="seconds"), name, "LINK_FAIL", None, None, None, None, None, None, ""))
        self.summary_lines.append("[OK] optimize ASM derleme denemesi tamamlandi.")

    def measure_exe(self, exe: Path) -> tuple[Optional[float], str, Optional[int]]:
        if not exe.exists():
            return None, "", None
        start = time.perf_counter_ns()
        proc = subprocess.run(str(exe), shell=True, capture_output=True, text=True, errors="replace")
        elapsed_ms = (time.perf_counter_ns() - start) / 1_000_000
        return elapsed_ms, (proc.stdout or "").strip(), proc.returncode

    def compare_optimized_outputs(self) -> None:
        exe_dir = self.base / "build" / "exe"
        seen = {r.test_name for r in self.opt_results}
        for opt_exe in sorted(exe_dir.glob("*_opt.exe")):
            name = opt_exe.stem[:-4]  # _opt sil
            if name in seen:
                continue
            orig_exe = exe_dir / f"{name}.exe"
            t_orig, out_orig, rc_orig = self.measure_exe(orig_exe)
            t_opt, out_opt, rc_opt = self.measure_exe(opt_exe)
            if t_orig is None or t_opt is None:
                self.opt_results.append(OptResult(self.stage, _dt.datetime.now().isoformat(timespec="seconds"), name, "MISSING_EXE", t_orig, t_opt, None, None, rc_orig, rc_opt, ""))
                continue
            match = out_orig == out_opt
            gain = ((t_orig - t_opt) / t_orig * 100.0) if t_orig and t_orig > 0 else None
            if not match:
                status = "OUTPUT_DIFF"
            elif gain is not None and gain > 0:
                status = "OPT_FASTER"
            else:
                status = "OPT_SLOWER_OR_EQUAL"
            diff_file = ""
            if not match:
                diff_path = self.opt_log_dir / f"diff_{safe_stem(name)}.txt"
                diff = difflib.unified_diff(
                    out_orig.splitlines(keepends=True),
                    out_opt.splitlines(keepends=True),
                    fromfile=f"{name}.exe",
                    tofile=f"{name}_opt.exe",
                )
                diff_path.write_text("".join(diff), encoding="utf-8")
                diff_file = str(diff_path.relative_to(self.base)).replace("/", "\\")
            self.opt_results.append(OptResult(
                self.stage,
                _dt.datetime.now().isoformat(timespec="seconds"),
                name,
                status,
                round(t_orig, 6),
                round(t_opt, 6),
                round(gain, 4) if gain is not None else None,
                match,
                rc_orig,
                rc_opt,
                diff_file,
            ))
        self.summary_lines.append("[OK] orijinal/optimize EXE karsilastirma tamamlandi.")

    def write_optimization_reports(self) -> None:
        if not self.opt_results:
            return
        csv_path = self.run_dir / "optimization_results.csv"
        write_dataclass_csv(csv_path, self.opt_results)

        db_path = self.report_dir / "uxm_perf_history.db"
        conn = sqlite3.connect(db_path)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS perf_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                stage INTEGER,
                test_name TEXT,
                timestamp TEXT,
                orig_ms REAL,
                opt_ms REAL,
                gain_perc REAL,
                status TEXT,
                output_match INTEGER,
                diff_file TEXT
            )
        """)
        for r in self.opt_results:
            conn.execute(
                "INSERT INTO perf_logs(stage,test_name,timestamp,orig_ms,opt_ms,gain_perc,status,output_match,diff_file) VALUES(?,?,?,?,?,?,?,?,?)",
                (r.stage, r.test_name, r.timestamp, r.orig_ms, r.opt_ms, r.gain_percent, r.status, None if r.output_match is None else int(r.output_match), r.diff_file),
            )
        conn.commit()
        conn.close()
        self.summary_lines.append(f"[OK] optimizer raporlari: {csv_path}, {db_path}")

    # -----------------------------
    # Raporlar ve arsivleme
    # -----------------------------
    def write_test_reports(self) -> None:
        current_csv = self.run_dir / "test_results.csv"
        write_dataclass_csv(current_csv, self.test_results)

        history_path = self.base / "test_history.csv"
        append_dataclass_csv(history_path, self.test_results)

        summary_path = self.base / "test_stats_summary.csv"
        self.write_stats_summary(history_path, summary_path)

        xlsx_path = self.run_dir / f"UXM_Stage_{self.stage}_Rapor.xlsx"
        self.try_write_xlsx(current_csv, summary_path, xlsx_path)

        self.summary_lines.append(f"[OK] test raporu: {current_csv}")
        self.summary_lines.append(f"[OK] gecmis/stat: {history_path}, {summary_path}")

    def write_stats_summary(self, history_path: Path, summary_path: Path) -> None:
        rows = read_csv_dicts(history_path)
        groups: dict[str, list[float]] = defaultdict(list)
        last_status: dict[str, str] = {}
        last_stage: dict[str, str] = {}
        for r in rows:
            name = r.get("test_path") or r.get("Test_Adi") or r.get("test_name") or ""
            try:
                val = float(r.get("duration_sec") or r.get("Son_Sure") or r.get("Sure_Sn") or 0)
            except ValueError:
                continue
            if name:
                groups[name].append(val)
                last_status[name] = r.get("status") or r.get("Durum") or ""
                last_stage[name] = str(r.get("stage") or "")

        with summary_path.open("w", encoding="utf-8", newline="") as f:
            fieldnames = ["Test_Adi", "Calistirma_Sayisi", "Ortalama_Sure", "Median_Sure", "Min_Sure", "Max_Sure", "Son_Sure", "Son_Status", "Son_Stage"]
            w = csv.DictWriter(f, fieldnames=fieldnames)
            w.writeheader()
            for name in sorted(groups):
                vals = groups[name]
                w.writerow({
                    "Test_Adi": name,
                    "Calistirma_Sayisi": len(vals),
                    "Ortalama_Sure": round(statistics.mean(vals), 6),
                    "Median_Sure": round(statistics.median(vals), 6),
                    "Min_Sure": round(min(vals), 6),
                    "Max_Sure": round(max(vals), 6),
                    "Son_Sure": round(vals[-1], 6),
                    "Son_Status": last_status.get(name, ""),
                    "Son_Stage": last_stage.get(name, ""),
                })

    def try_write_xlsx(self, current_csv: Path, summary_csv: Path, xlsx_path: Path) -> None:
        if self.args.no_excel:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except Exception:
            self.summary_lines.append("[WARN] openpyxl yok; xlsx raporu atlandi.")
            return
        wb = Workbook()
        default = wb.active
        wb.remove(default)
        for sheet_name, csv_path in [
            ("Current_Run", current_csv),
            ("Stats_Summary", summary_csv),
            ("Optimization", self.run_dir / "optimization_results.csv"),
        ]:
            if not csv_path.exists():
                continue
            ws = wb.create_sheet(sheet_name[:31])
            for row_idx, row in enumerate(read_csv_rows(csv_path), 1):
                ws.append(row)
                if row_idx == 1:
                    for cell in ws[row_idx]:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill("solid", fgColor="366092")
                        cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                width = min(max(len(str(c.value)) if c.value is not None else 0 for c in col) + 2, 70)
                ws.column_dimensions[col[0].column_letter].width = width
        if wb.sheetnames:
            wb.save(xlsx_path)
            self.summary_lines.append(f"[OK] xlsx raporu: {xlsx_path}")

    def archive_build(self) -> Path:
        build = self.base / "build"
        if not build.exists():
            self.summary_lines.append("[WARN] build klasoru yok; arsivleme atlandi.")
            return build
        target = self.base / f"build stage {self.stage}"
        if target.exists():
            target = self.base / f"build stage {self.stage}_{self.run_stamp}"
        shutil.copytree(build, target)
        self.summary_lines.append(f"[OK] final build klasoru: {target.name}")
        return target

    def write_final_summary(self, final_build: Path, ok: bool) -> None:
        elapsed = time.perf_counter() - self.session_start
        passed = sum(1 for r in self.test_results if r.status == "BASARILI")
        failed = sum(1 for r in self.test_results if r.status != "BASARILI")
        opt_bad = sum(1 for r in self.opt_results if r.status in {"OUTPUT_DIFF", "NASM_FAIL", "LINK_FAIL", "MISSING_EXE"})
        slowest = sorted(self.test_results, key=lambda r: r.duration_sec, reverse=True)[:10]

        md = []
        md.append(f"# UXM Stage {self.stage} Run Summary")
        md.append("")
        md.append(f"- Tarih: {self.now.isoformat(timespec='seconds')}")
        md.append(f"- Durum: {'OK' if ok and failed == 0 and opt_bad == 0 else 'KONTROL_GEREKLI'}")
        md.append(f"- Sonuc log: `{self.result_log.name}`")
        md.append(f"- Run klasoru: `{self.run_dir.relative_to(self.base)}`")
        md.append(f"- Final build klasoru: `{final_build.name}`")
        md.append(f"- Toplam sure: {elapsed:.2f} sn")
        md.append(f"- Compiler build sure: {self.build_duration:.3f} sn")
        md.append(f"- Test: {passed} basarili / {failed} basarisiz / toplam {len(self.test_results)}")
        md.append(f"- Optimizer kritik durum: {opt_bad}")
        md.append("")
        md.append("## Faz Ozeti")
        for line in self.summary_lines:
            md.append(f"- {line}")
        md.append("")
        md.append("## En Yavas 10 Test")
        md.append("| Test | Sure sn | Durum |")
        md.append("|---|---:|---|")
        for r in slowest:
            md.append(f"| `{r.test_path}` | {r.duration_sec:.3f} | {r.status} |")
        if self.opt_results:
            md.append("")
            md.append("## Optimizer Ozeti")
            counts = Counter(r.status for r in self.opt_results)
            for k, v in sorted(counts.items()):
                md.append(f"- {k}: {v}")
        md.append("")
        summary_path = self.run_dir / "STAGE_RUN_SUMMARY.md"
        summary_path.write_text("\n".join(md) + "\n", encoding="utf-8")
        self.append_session_log(f"END_SESSION@{_dt.datetime.now().strftime('%Y-%m-%d')}@{_dt.datetime.now().strftime('%H:%M:%S')}@SECONDS@{elapsed:.6f}")
        self.summary_lines.append(f"[OK] final ozet: {summary_path}")

        state = {
            "last_completed_stage": self.stage if failed == 0 else self.stage - 1,
            "current_stage": self.stage + 1 if failed == 0 else self.stage,
            "last_run": self.now.isoformat(timespec="seconds"),
            "last_result_log": self.result_log.name,
            "last_final_build": final_build.name,
            "last_status": "OK" if failed == 0 else "FAIL",
        }
        (self.base / STAGE_STATE_FILE).write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")

    def run(self) -> int:
        print(f"UXM Stage Runner basladi: stage={self.stage}, log={self.result_log.name}")
        self.init_log()
        if not self.build_compiler() and not self.args.continue_on_fail:
            self.write_test_reports()
            final = self.archive_build()
            self.write_final_summary(final, ok=False)
            return 1
        smoke_ok = self.run_smoke_phase()
        if not smoke_ok and not self.args.continue_on_fail:
            self.write_test_reports()
            final = self.archive_build()
            self.write_final_summary(final, ok=False)
            return 1
        tests_ok = self.run_full_tests()
        self.write_test_reports()
        self.optimizer_phase()
        # Optimizer raporu xlsx'e girsin diye tekrar dene.
        self.try_write_xlsx(self.run_dir / "test_results.csv", self.base / "test_stats_summary.csv", self.run_dir / f"UXM_Stage_{self.stage}_Rapor.xlsx")
        final = self.archive_build()
        ok = bool(smoke_ok and tests_ok)
        self.write_final_summary(final, ok=ok)
        print(f"Bitti. Final build klasoru: {final.name}")
        print(f"Ozet: {self.run_dir / 'STAGE_RUN_SUMMARY.md'}")
        return 0 if ok else 1


# -----------------------------
# ASM kurallari
# -----------------------------

def clean_instr(line: str) -> Optional[str]:
    s = line.split(";", 1)[0].strip()
    if not s or s.endswith(":"):
        return None
    return s


def apply_safe_asm_rules(lines: list[str]) -> list[str]:
    """Derlenebilirligi bozma riski dusuk, temkinli kurallar.
    Kritik ABI/flag etkisi olan kurallari sadece yorum olarak isaretler.
    """
    out: list[str] = []
    i = 0
    while i < len(lines):
        raw = lines[i].rstrip("\n")
        line = raw.strip()
        low = line.lower()
        # mov reg, 0 -> xor reg, reg: x64 icin guvenli sayilir; flags degisir ama mov da flags degistirmez.
        # Bu yuzden her yerde otomatik degistirmek riskli olabilir; sadece cok bilinen scratch registerlarda uygula.
        m = re.match(r"mov\s+(r(?:ax|bx|cx|dx|8|9|10|11)),\s*0\s*$", low)
        if m:
            reg = m.group(1)
            out.append(f"xor {reg}, {reg} ; [UXM_OPT: mov reg,0]")
            i += 1
            continue
        # jmp hemen sonraki label'a gidiyorsa sil.
        if low.startswith("jmp ") and i + 1 < len(lines):
            label = line.split()[-1]
            if lines[i + 1].strip() == f"{label}:":
                out.append(f"; [UXM_OPT: jmp-to-next removed] {raw}")
                i += 1
                continue
        out.append(raw)
        i += 1
    return out


# -----------------------------
# CSV / quoting yardimcilari
# -----------------------------

def quote_win(s: str) -> str:
    # Komut argumanlarini Windows shell icin guvenli quote et.
    s = str(s)
    if len(s) >= 2 and s[0] == '"' and s[-1] == '"':
        return s
    return '"' + s.replace('"', '\\"') + '"'


def safe_stem(s: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", s)


def write_dataclass_csv(path: Path, rows: list) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    fields = list(asdict(rows[0]).keys())
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for row in rows:
            w.writerow(asdict(row))


def append_dataclass_csv(path: Path, rows: list) -> None:
    if not rows:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = list(asdict(rows[0]).keys())
    write_header = (not path.exists()) or path.stat().st_size == 0 or not has_header(path, fields)
    mode = "a" if not write_header else "w"
    # Eski bozuk/headersiz history varsa onu yedekle, yeni temiz format ac.
    if path.exists() and path.stat().st_size > 0 and write_header:
        backup = path.with_suffix(path.suffix + f".bak_{_dt.datetime.now().strftime('%Y%m%d_%H%M%S')}")
        shutil.copy2(path, backup)
    with path.open(mode, encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        if write_header:
            w.writeheader()
        for row in rows:
            w.writerow(asdict(row))


def has_header(path: Path, expected_fields: list[str]) -> bool:
    try:
        with path.open("r", encoding="utf-8", newline="") as f:
            first = f.readline().strip("\ufeff\r\n")
        if not first:
            return False
        cols = [c.strip() for c in first.split(",")]
        return all(x in cols for x in ["stage", "test_path", "duration_sec", "status"])
    except Exception:
        return False


def read_csv_dicts(path: Path) -> list[dict]:
    if not path.exists() or path.stat().st_size == 0:
        return []
    try:
        with path.open("r", encoding="utf-8", newline="") as f:
            return list(csv.DictReader(f))
    except Exception:
        return []


def read_csv_rows(path: Path) -> Iterable[list]:
    if not path.exists() or path.stat().st_size == 0:
        return []
    with path.open("r", encoding="utf-8", newline="") as f:
        return list(csv.reader(f))


# -----------------------------
# CLI
# -----------------------------

def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="UXM stage smoke/full-test/optimizer runner")
    p.add_argument("--stage", default="auto", help="Stage numarasi veya auto. Ornek: --stage 12")
    p.add_argument("--fbc", default=None, help="FreeBASIC fbc.exe yolu. Bos kalirsa UXM_FBC/FBC/env/path aranir.")
    p.add_argument("--nasm", default=None, help="NASM komutu veya yolu. Bos kalirsa UXM_NASM/nasm kullanilir.")
    p.add_argument("--skip-opt", action="store_true", help="ASM optimizer fazini atla.")
    p.add_argument("--all-smoke", action="store_true", help="Bulunan tum smoke bat dosyalarini calistir; varsayilan sadece en guncel smoke.")
    p.add_argument("--continue-on-fail", action="store_true", help="Smoke/test hata verse bile sonraki fazlara devam et.")
    p.add_argument("--include-tmp", action="store_true", help="_tmp*.uxm debug testlerini de dahil et.")
    p.add_argument("--no-excel", action="store_true", help="XLSX rapor denemesini kapat.")
    return p


def main(argv: Optional[list[str]] = None) -> int:
    args = build_arg_parser().parse_args(argv)
    runner = StageRunner(args)
    return runner.run()


if __name__ == "__main__":
    raise SystemExit(main())
