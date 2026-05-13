import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def format_excel_worksheet(ws):
    """Excel sekmesini profesyonel formatta süsler."""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border

def run_final_analysis():
    file_path = 'test_history.csv'
    
    if not os.path.exists(file_path):
        print(f"Hata: {file_path} bulunamadı!")
        return

    # Dosyayı oku
    df = pd.read_csv(file_path)
    
    # HATA ÇÖZÜMÜ: Sütun isimlerine değil sıralarına (0,1,2,3) güveniyoruz
    # 1. sütun Test Adı, 2. sütun Süre, 3. sütun Derleme Süresi
    col_test_name = df.columns[1]
    col_duration = df.columns[2]
    col_build_time = df.columns[3]

    # --- SEKME 1: GENEL İSTATİSTİKLER ---
    summary_stats = df.groupby(col_test_name).agg({
        col_duration: ['count', 'mean', 'std', 'min', 'max', 'last']
    }).reset_index()
    summary_stats.columns = ['Test_Adi', 'Kosma_Sayisi', 'Ort_Sure', 'Std_Sapma', 'En_Hizli', 'En_Yavas', 'Son_Sure']

    # --- SEKME 2: PERFORMANS TRENDİ (İyileşme Oranı) ---
    trend_list = []
    for test in df[col_test_name].unique():
        test_data = df[df[col_test_name] == test]
        if len(test_data) > 1:
            first = test_data[col_duration].iloc[0]
            last = test_data[col_duration].iloc[-1]
            improvement = ((first - last) / first) * 100 if first != 0 else 0
            trend_list.append({
                'Test_Adi': test, 
                'Ilk_Sure': round(first, 4), 
                'Son_Sure': round(last, 4), 
                'Kazanc_%': round(improvement, 2)
            })
    df_trend = pd.DataFrame(trend_list)

    # --- EXCEL ÜRETİMİ ---
    wb = Workbook()
    
    # 1. Sekme: Genel İstatistikler
    ws1 = wb.active
    ws1.title = "Genel İstatistikler"
    for r in dataframe_to_rows(summary_stats, index=False, header=True):
        ws1.append(r)
    format_excel_worksheet(ws1)

    # 2. Sekme: Performans Trendi
    ws2 = wb.create_sheet("Performans Trendi")
    if not df_trend.empty:
        for r in dataframe_to_rows(df_trend, index=False, header=True):
            ws2.append(r)
    format_excel_worksheet(ws2)

    # 3. Sekme: Detaylı Geçmiş
    ws3 = wb.create_sheet("Tum Kosu Detaylari")
    for r in dataframe_to_rows(df.sort_values([col_test_name, df.columns[0]]), index=False, header=True):
        ws3.append(r)
    format_excel_worksheet(ws3)

    excel_name = "UXM_Performans_Analiz_Raporu_Final.xlsx"
    wb.save(excel_name)

    # --- TERMİNAL ÇIKTISI (Bulgular) ---
    print("\n" + "="*60)
    print("   UXM PERFORMANS ANALİZİ - KRİTİK BULGULAR")
    print("="*60)
    
    if not df_trend.empty:
        top_gain = df_trend.sort_values('Kazanc_%', ascending=False).head(3)
        print("\n[+] EN ÇOK HIZLANAN TESTLER:")
        for _, row in top_gain.iterrows():
            print(f" - {row['Test_Adi']}: %{row['Kazanc_%']} iyileşme")

    unstable = summary_stats.sort_values('Std_Sapma', ascending=False).head(2)
    print("\n[!] EN YÜKSEK VARYANS (İstikrarsızlık):")
    for _, row in unstable.iterrows():
        print(f" - {row['Test_Adi']}: Sapma {round(row['Std_Sapma'], 4)}")

    print(f"\n[*] Ortalama Derleme Süresi: {round(df[col_build_time].mean(), 2)} sn")
    print(f"\n3 Sekmeli Excel Raporu Hazır: {excel_name}")
    print("="*60)

if __name__ == "__main__":
    run_final_analysis()