# UXM PERFORMANS ANALİZİ - LEGACY CSV UYUMLU DÜZELTİLMİŞ SÜRÜM
# Girdi: test_history.csv 4 kolon legacy formatı
# Kolonlar: Tarih, Test_Adi, Son_Sure, Derleme_Sn

import os
import csv
import datetime
import statistics

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
except Exception as exc:
    pd = None
    Workbook = None
    _IMPORT_ERROR = exc

LEGACY_HEADER = ["Tarih", "Test_Adi", "Son_Sure", "Derleme_Sn"]


def load_legacy_history(path="test_history.csv"):
    rows = []
    if not os.path.exists(path):
        raise FileNotFoundError(path)
    with open(path, "r", encoding="utf-8-sig", errors="replace", newline="") as f:
        for row in csv.reader(f):
            if not row or row == LEGACY_HEADER:
                continue
            if len(row) < 4:
                continue
            # Yeni runner'ın 10 kolonlu satırlarını kesinlikle alma.
            if len(row) != 4:
                continue
            try:
                sure = float(row[2].replace(',', '.'))
                build = float(row[3].replace(',', '.'))
            except Exception:
                continue
            rows.append({"Tarih": row[0], "Test_Adi": row[1], "Son_Sure": sure, "Derleme_Sn": build})
    return rows


def format_excel_worksheet(ws):
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 55)


def run_final_analysis():
    if pd is None:
        print(f"Hata: pandas/openpyxl yüklenemedi: {_IMPORT_ERROR}")
        return 1
    try:
        rows = load_legacy_history("test_history.csv")
    except FileNotFoundError:
        print("Hata: test_history.csv bulunamadı!")
        return 1
    if not rows:
        print("Hata: test_history.csv içinde okunabilir 4 kolon legacy veri yok.")
        return 1

    df = pd.DataFrame(rows)
    summary_stats = df.groupby("Test_Adi").agg(
        Kosma_Sayisi=("Son_Sure", "count"),
        Ort_Sure=("Son_Sure", "mean"),
        Std_Sapma=("Son_Sure", "std"),
        En_Hizli=("Son_Sure", "min"),
        En_Yavas=("Son_Sure", "max"),
        Son_Sure=("Son_Sure", "last"),
    ).reset_index()
    summary_stats["Std_Sapma"] = summary_stats["Std_Sapma"].fillna(0)

    trend_list = []
    for test, test_data in df.groupby("Test_Adi", sort=False):
        if len(test_data) > 1:
            first = float(test_data["Son_Sure"].iloc[0])
            last = float(test_data["Son_Sure"].iloc[-1])
            improvement = ((first - last) / first) * 100 if first else 0
            trend_list.append({"Test_Adi": test, "Ilk_Sure": round(first, 4), "Son_Sure": round(last, 4), "Kazanc_%": round(improvement, 2)})
    df_trend = pd.DataFrame(trend_list)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Genel Istatistikler"
    for r in dataframe_to_rows(summary_stats.round(4), index=False, header=True):
        ws1.append(r)
    format_excel_worksheet(ws1)

    ws2 = wb.create_sheet("Performans Trendi")
    if not df_trend.empty:
        for r in dataframe_to_rows(df_trend, index=False, header=True):
            ws2.append(r)
    else:
        ws2.append(["Bilgi"]); ws2.append(["Trend için her test en az iki kez çalışmalı."])
    format_excel_worksheet(ws2)

    ws3 = wb.create_sheet("Tum Kosu Detaylari")
    for r in dataframe_to_rows(df.sort_values(["Test_Adi", "Tarih"]), index=False, header=True):
        ws3.append(r)
    format_excel_worksheet(ws3)

    excel_name = "UXM_Performans_Analiz_Raporu_Final.xlsx"
    try:
        wb.save(excel_name)
    except PermissionError:
        print(f"[HATA] {excel_name} açık. Excel'i kapatıp tekrar çalıştır.")
        return 2

    print("=" * 60)
    print("UXM PERFORMANS ANALİZİ - KRİTİK BULGULAR")
    print("=" * 60)
    if not df_trend.empty:
        for _, row in df_trend.sort_values('Kazanc_%', ascending=False).head(3).iterrows():
            print(f"[HIZLANMA] {row['Test_Adi']}: %{row['Kazanc_%']}")
    for _, row in summary_stats.sort_values('Std_Sapma', ascending=False).head(3).iterrows():
        print(f"[VARYANS] {row['Test_Adi']}: {row['Std_Sapma']:.4f}")
    print(f"Ortalama derleme süresi: {df['Derleme_Sn'].mean():.2f} sn")
    print(f"Excel raporu: {excel_name}")
    return 0


if __name__ == "__main__":
    raise SystemExit(run_final_analysis())
