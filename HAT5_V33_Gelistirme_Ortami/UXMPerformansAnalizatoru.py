#UXM PERFORMANS ANALİZİ


""" 
==================================================
   UXM PERFORMANS ANALİZİ - İLGİNÇ BULGULAR
==================================================

[+] EN ÇOK İYİLEŞEN TESTLER (Hızlanma):
 - test_str_find: %12.5 kazanç
 - test_v33_tensor: %8.89 kazanç
 - test_fp01: %6.25 kazanç

[!] EN YÜKSEK VARYANS (İstikrarsız Süreler):
 - test_fp01: Std Sapma 0.0645
 - test_str_find: Std Sapma 0.0645
 - test_fp02: Std Sapma 0.0457

[*] Ortalama Derleme Süresi: 4.54 saniye
==================================================

Detaylı Excel Raporu Hazır: UXM_Performans_Analiz_Raporu.xlsx 
 bu program stsx,py ile derlenen test runnerin urettigi bilgileri analiz eder,
 istatistiksel ozetler sunar ve xlsx dosya ile bana daha baska analiz yapma imkani yaratan 
rapor olusturucudur.
"""


import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def format_excel_worksheet(ws):
    """Excel sekmesini profesyonel formatta süsler."""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border

def run_final_analysis():
    file_path = 'test_history.csv'
    
    if not os.path.exists(file_path):
        print(f"Hata: {file_path} bulunamadı!")
        return

    # Dosyayı oku
    df = pd.read_csv(file_path)
    
    # HATA DÜZELTME: Süre sütununu sayıya çevir (Hatalı verileri NaN yap ve temizle)
    col_duration = df.columns[2] # Süre sütunu
    df[col_duration] = pd.to_numeric(df[col_duration].astype(str).str.replace(',', '.'), errors='coerce')
    df = df.dropna(subset=[col_duration]) # Sayıya çevrilemeyen boş satırları temizle
    
    # HATA ÇÖZÜMÜ: Sütun isimlerine değil sıralarına (0,1,2,3) güveniyoruz
    # 1. sütun Test Adı, 2. sütun Süre, 3. sütun Derleme Süresi
    col_test_name = df.columns[1]
    col_duration = df.columns[2]
    col_build_time = df.columns[3]

    # --- SEKME 1: GENEL İSTATİSTİKLER ---
    summary_stats = df.groupby(col_test_name).agg({
        col_duration: ['count', 'mean', 'std', 'min', 'max', 'last']
    }).reset_index()
    summary_stats.columns = ['Test_Adi', 'Kosma_Sayisi', 'Ort_Sure', 'Std_Sapma', 'En_Hizli', 'En_Yavas', 'Son_Sure']

    # --- SEKME 2: PERFORMANS TRENDİ (İyileşme Oranı) ---
    trend_list = []
    for test in df[col_test_name].unique():
        test_data = df[df[col_test_name] == test]
        if len(test_data) > 1:
            first = test_data[col_duration].iloc[0]
            last = test_data[col_duration].iloc[-1]
            improvement = ((first - last) / first) * 100 if first != 0 else 0
            trend_list.append({
                'Test_Adi': test, 
                'Ilk_Sure': round(first, 4), 
                'Son_Sure': round(last, 4), 
                'Kazanc_%': round(improvement, 2)
            })
    df_trend = pd.DataFrame(trend_list)

    # --- EXCEL ÜRETİMİ ---
    try:
        wb.save(excel_name)
        print(f"\n3 Sekmeli Excel Raporu Hazır: {excel_name}")
    except PermissionError:
        print(f"\n[!!!] HATA: {excel_name} dosyası şu an açık!")
        print("Lütfen Excel dosyasını kapatın ve programı tekrar çalıştırın.")
    
    # 1. Sekme: Genel İstatistikler
    ws1 = wb.active
    ws1.title = "Genel İstatistikler"
    for r in dataframe_to_rows(summary_stats, index=False, header=True):
        ws1.append(r)
    format_excel_worksheet(ws1)

    # 2. Sekme: Performans Trendi
    ws2 = wb.create_sheet("Performans Trendi")
    if not df_trend.empty:
        for r in dataframe_to_rows(df_trend, index=False, header=True):
            ws2.append(r)
    format_excel_worksheet(ws2)

    # 3. Sekme: Detaylı Geçmiş
    ws3 = wb.create_sheet("Tum Kosu Detaylari")
    for r in dataframe_to_rows(df.sort_values([col_test_name, df.columns[0]]), index=False, header=True):
        ws3.append(r)
    format_excel_worksheet(ws3)

    excel_name = "UXM_Performans_Analiz_Raporu_Final.xlsx"
    try:
        wb.save(excel_name)
        print(f"\n3 Sekmeli Excel Raporu Hazır: {excel_name}")
    except PermissionError:
        print(f"\n[!!!] HATA: {excel_name} dosyası şu an açık!")
        print("Lütfen Excel dosyasını kapatın ve programı tekrar çalıştırın.")

    # --- TERMİNAL ÇIKTISI (Bulgular) ---
    print("\n" + "="*60)
    print("   UXM PERFORMANS ANALİZİ - KRİTİK BULGULAR")
    print("="*60)
    
    if not df_trend.empty:
        top_gain = df_trend.sort_values('Kazanc_%', ascending=False).head(3)
        print("\n[+] EN ÇOK HIZLANAN TESTLER:")
        for _, row in top_gain.iterrows():
            print(f" - {row['Test_Adi']}: %{row['Kazanc_%']} iyileşme")

    unstable = summary_stats.sort_values('Std_Sapma', ascending=False).head(2)
    print("\n[!] EN YÜKSEK VARYANS (İstikrarsızlık):")
    for _, row in unstable.iterrows():
        print(f" - {row['Test_Adi']}: Sapma {round(row['Std_Sapma'], 4)}")

    print(f"\n[*] Ortalama Derleme Süresi: {round(df[col_build_time].mean(), 2)} sn")
    print(f"\n3 Sekmeli Excel Raporu Hazır: {excel_name}")
    print("="*60)

if __name__ == "__main__":
    run_final_analysis()



















# import pandas as pd
# import os
# from openpyxl import Workbook
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# from openpyxl.utils.dataframe import dataframe_to_rows

# def run_smart_analysis():
#     file_path = 'test_history.csv'
    
#     if not os.path.exists(file_path):
#         print(f"Hata: {file_path} dosyası bulunamadı!")
#         return

#     # Dosyayı oku (Başlıkları otomatik tanıması için)
#     df = pd.read_csv(file_path)
    
#     # SÜTUN TESPİTİ: Sütun isimleri ne olursa olsun pozisyona göre alıyoruz
#     # 0: Tarih, 1: Test Adı, 2: Süre, 3: Derleme
#     col_test = df.columns[1]
#     col_sure = df.columns[2]
#     col_build = df.columns[3]
    
#     print(f"Analiz ediliyor: {col_test} sütunu baz alınıyor...")

#     # 1. Genel İstatistikler
#     summary_stats = df.groupby(col_test).agg({
#         col_sure: ['count', 'mean', 'std', 'min', 'max', 'last']
#     }).reset_index()
#     summary_stats.columns = ['Test_Adi', 'Kosma_Sayisi', 'Ort_Sure', 'Std_Sapma', 'En_Hizli', 'En_Yavas', 'Son_Sure']

#     # 2. Trend ve İyileşme Oranı
#     trend_list = []
#     for test in df[col_test].unique():
#         test_data = df[df[col_test] == test]
#         if len(test_data) > 1:
#             first = test_data[col_sure].iloc[0]
#             last = test_data[col_sure].iloc[-1]
#             improvement = ((first - last) / first) * 100 if first != 0 else 0
#             trend_list.append({'Test_Adi': test, 'Ilk_Sure': first, 'Son_Sure': last, 'Kazanc_%': round(improvement, 2)})
    
#     df_trend = pd.DataFrame(trend_list)

#     # --- EXCEL OLUŞTURMA ---
#     wb = Workbook()
#     ws1 = wb.active
#     ws1.title = "Istatistikler"

#     # Verileri Excel'e yaz
#     for r in dataframe_to_rows(summary_stats, index=False, header=True):
#         ws1.append(r)

#     excel_name = "UXM_Performans_Raporu_Final.xlsx"
#     wb.save(excel_name)

#     # --- TERMİNAL ÇIKTISI ---
#     print("\n" + "="*50)
#     print("   UXM PERFORMANS KRİTİK BULGULAR")
#     print("="*50)
    
#     if not df_trend.empty:
#         best = df_trend.sort_values('Kazanc_%', ascending=False).head(1)
#         print(f"\n[+] EN ÇOK HIZLANAN: {best['Test_Adi'].values[0]} (%{best['Kazanc_%'].values[0]} kazanç)")
    
#     print(f"\n[*] Ortalama Derleme: {round(df[col_build].mean(), 2)} sn")
#     print(f"\nExcel Raporu Oluştu: {excel_name}")
#     print("="*50)

# if __name__ == "__main__":
#     run_smart_analysis()


# birinci program csv dosyadaki test adi  sutununu ariyordu ama bulamadi bende ileride incelerim diye burada rem haline getirdim.
# import pandas as pd
# import numpy as np
# from openpyxl import Workbook
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# from openpyxl.utils.dataframe import dataframe_to_rows

# # Veri dosyasını yükleyelim (test_history.csv varsa onu, yoksa elimizdeki CSV'yi simüle edelim)
# # Kullanıcı tarafından paylaşılan verileri birleştirip analiz edeceğiz.
# file_path = 'test_history.csv'

# # Eğer dosya yoksa, kullanıcının paylaştığı snippet'lardan bir veri çerçevesi oluşturalım (simülasyon)
# try:
#     df = pd.read_csv(file_path)
# except:
#     # Simülasyon verisi (kullanıcı snippet'larına dayalı)
#     data = {
#         'Tarih': ['2026-05-08 08:20']*10 + ['2026-05-08 22:57']*10,
#         'Test_Adi': ['test_fp01', 'test_fp02', 'test_math01', 'test_v33_tensor', 'test_str_find']*4,
#         'Son_Sure': [2.4, 1.78, 1.7, 0.9, 1.2, 2.35, 1.75, 1.68, 0.88, 1.15, 2.30, 1.70, 1.65, 0.85, 1.10, 2.25, 1.68, 1.60, 0.82, 1.05],
#         'Derleme_Sn': [4.54]*20
#     }
#     df = pd.DataFrame(data)

# # --- ANALİZ BÖLÜMÜ ---

# # 1. Genel İstatistikler
# summary_stats = df.groupby('Test_Adi').agg({
#     'Son_Sure': ['count', 'mean', 'std', 'min', 'max', 'last']
# }).reset_index()
# summary_stats.columns = ['Test_Adi', 'Kosma_Sayisi', 'Ort_Sure', 'Std_Sapma', 'En_Hizli', 'En_Yavas', 'Son_Sure']

# # 2. İyileşme Oranı (Trend)
# # Her testin ilk ve son koşusu arasındaki fark
# trend_df = []
# for test in df['Test_Adi'].unique():
#     test_data = df[df['Test_Adi'] == test].sort_values('Tarih')
#     if len(test_data) > 1:
#         first = test_data['Son_Sure'].iloc[0]
#         last = test_data['Son_Sure'].iloc[-1]
#         improvement = ((first - last) / first) * 100
#         trend_df.append({'Test_Adi': test, 'Ilk_Sure': first, 'Son_Sure': last, 'Iyilestirme_%': round(improvement, 2)})

# df_trend = pd.DataFrame(trend_df)

# # --- EXCEL OLUŞTURMA ---
# wb = Workbook()
# ws_summary = wb.active
# ws_summary.title = "Genel İstatistikler"

# # Stil Tanımlamaları
# header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
# header_font = Font(color="FFFFFF", bold=True)
# thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# def format_excel(ws):
#     for row in ws.iter_rows(min_row=1, max_row=1):
#         for cell in row:
#             cell.fill = header_fill
#             cell.font = header_font
#             cell.alignment = Alignment(horizontal="center")
#     for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
#         for cell in row:
#             cell.border = thin_border

# # Sayfa 1: Genel Özet
# for r in dataframe_to_rows(summary_stats, index=False, header=True):
#     ws_summary.append(r)
# format_excel(ws_summary)

# # Sayfa 2: Trend ve İyileşme
# ws_trend = wb.create_sheet("Performans Trendi")
# for r in dataframe_to_rows(df_trend, index=False, header=True):
#     ws_trend.append(r)
# format_excel(ws_trend)

# # Sayfa 3: Detaylı Geçmiş (Her dosya için ayrı seçilebilir yapı simülasyonu)
# ws_history = wb.create_sheet("Tüm Koşu Detayları")
# for r in dataframe_to_rows(df.sort_values(['Test_Adi', 'Tarih']), index=False, header=True):
#     ws_history.append(r)
# format_excel(ws_history)

# excel_file = "UXM_Performans_Analiz_Raporu.xlsx"
# wb.save(excel_file)

# # --- TERMİNAL ÇIKTISI ---
# print("\n" + "="*50)
# print("   UXM PERFORMANS ANALİZİ - İLGİNÇ BULGULAR")
# print("="*50)

# # En çok iyileşen 3 test
# if not df_trend.empty:
#     top_improvement = df_trend.sort_values('Iyilestirme_%', ascending=False).head(3)
#     print("\n[+] EN ÇOK İYİLEŞEN TESTLER (Hızlanma):")
#     for _, row in top_improvement.iterrows():
#         print(f" - {row['Test_Adi']}: %{row['Iyilestirme_%']} kazanç")

# # En istikrarsız testler (Std Sapma yüksek olanlar)
# unstable = summary_stats.sort_values('Std_Sapma', ascending=False).head(3)
# print("\n[!] EN YÜKSEK VARYANS (İstikrarsız Süreler):")
# for _, row in unstable.iterrows():
#     print(f" - {row['Test_Adi']}: Std Sapma {round(row['Std_Sapma'], 4)}")

# # Genel ortalama derleme süresi
# avg_build = df['Derleme_Sn'].mean()
# print(f"\n[*] Ortalama Derleme Süresi: {round(avg_build, 2)} saniye")
# print("="*50)
# print(f"\nDetaylı Excel Raporu Hazır: {excel_file}")