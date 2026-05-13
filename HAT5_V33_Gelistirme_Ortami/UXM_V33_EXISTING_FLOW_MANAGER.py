# -*- coding: utf-8 -*-
"""
UXM V33 Existing Flow Manager
-----------------------------
Amaç:
- Mete abi'nin mevcut build/test/optimizer dosyalarını yeniden üretmeden sıraya koymak.
- Smoke testleri önce çalıştırmak, full testleri build_one_native.bat ile ölçmek.
- test_history.csv ve test_stats_summary.csv geçmişini koruyarak süre istatistiği üretmek.
- XLSX raporlarında kontrol karakteri yüzünden çökmeden güvenli yazmak.
- run_opt.bat içindeki hatalı/missing dosya durumlarını raporlamak; mevcut Python sınıflarını doğru sırayla çağırmak.
- İş bitince build klasörünü ve eski rapor/logları güvenli şekilde emekli klasörüne taşıyabilmek.

Kullanım örnekleri:
    python UXM_V33_EXISTING_FLOW_MANAGER.py run --stage 12
    python UXM_V33_EXISTING_FLOW_MANAGER.py run --stage 12 --skip-smoke
    python UXM_V33_EXISTING_FLOW_MANAGER.py opt --stage 12
    python UXM_V33_EXISTING_FLOW_MANAGER.py audit --stage 12
    python UXM_V33_EXISTING_FLOW_MANAGER.py toparla --stage 12 --dry-run
    python UXM_V33_EXISTING_FLOW_MANAGER.py toparla --stage 12 --apply
    python UXM_V33_EXISTING_FLOW_MANAGER.py pause-patch --dry-run
    python UXM_V33_EXISTING_FLOW_MANAGER.py pause-patch --apply
"""

from __future__ import annotations

import argparse
import csv
import datetime as _dt
import difflib
import json
import math
import os
import re
import shutil
import statistics
import subprocess
import sys
import time
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

ILLEGAL_XLSX_CHARS_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
SONUC_RE = re.compile(r"^sonuc(\d+)\.txt$", re.IGNORECASE)
RUN_EXPECT_RE = re.compile(r"^\s*call\s+:RUN_EXPECT\s+(.+?)\s+([^\s]+)\s+([^\s]+)\s*$", re.IGNORECASE)

TEST_DIR_PRIORITY = [
    Path("uxm/tests/fp"),
    Path("uxm/tests/math"),
    Path("uxm/tests/matrix"),
    Path("uxm/tests/native"),
    Path("uxm/tests/v33"),
]


def now_stamp() -> str:
    return _dt.datetime.now().strftime("%Y%m%d_%H%M%S")


def now_human() -> str:
    return _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def relpath(path: Path, root: Path) -> str:
    try:
        return str(path.resolve().relative_to(root.resolve())).replace("/", "\\")
    except Exception:
        return str(path).replace("/", "\\")


def decode_bytes(data: bytes) -> str:
    for enc in ("utf-8", "cp1254", "cp1252", "latin-1"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            pass
    return data.decode("utf-8", errors="replace")


def xlsx_safe(value):
    if value is None:
        return ""
    if isinstance(value, (int, float, bool)):
        return value
    return ILLEGAL_XLSX_CHARS_RE.sub(lambda m: "\\x%02X" % ord(m.group(0)), str(value))


def compact_text(text: str) -> str:
    # Beklenen çıktı karşılaştırması için boşlukları ve kontrol karakterlerini sadeleştirir.
    out = []
    for ch in text:
        if ch.isspace():
            continue
        if ord(ch) < 32:
            continue
        out.append(ch)
    return "".join(out)


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def append_csv(path: Path, header: Sequence[str], row: Sequence, write_header_if_new: bool = True) -> None:
    ensure_dir(path.parent)
    exists = path.exists() and path.stat().st_size > 0
    with path.open("a", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        if write_header_if_new and not exists:
            w.writerow(header)
        w.writerow([xlsx_safe(x) for x in row])


def write_csv(path: Path, header: Sequence[str], rows: Iterable[Sequence]) -> None:
    ensure_dir(path.parent)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        for row in rows:
            w.writerow([xlsx_safe(x) for x in row])


def write_text(path: Path, text: str) -> None:
    ensure_dir(path.parent)
    path.write_text(text, encoding="utf-8")


def run_shell(cmd: str, cwd: Path, timeout: Optional[int] = None) -> Tuple[int, str, str, float]:
    start = time.perf_counter()
    try:
        p = subprocess.run(
            cmd,
            cwd=str(cwd),
            shell=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            timeout=timeout,
        )
        elapsed = time.perf_counter() - start
        return p.returncode, decode_bytes(p.stdout), decode_bytes(p.stderr), elapsed
    except subprocess.TimeoutExpired as exc:
        elapsed = time.perf_counter() - start
        out = decode_bytes(exc.stdout or b"")
        err = decode_bytes(exc.stderr or b"") + f"\n[TIMEOUT] {timeout} sn siniri asildi."
        return 124, out, err, elapsed


def next_sonuc_number(root: Path) -> int:
    nums = []
    for p in root.glob("sonuc*.txt"):
        m = SONUC_RE.match(p.name)
        if m:
            nums.append(int(m.group(1)))
    return (max(nums) + 1) if nums else 1


def detect_stage_from_build_names(root: Path) -> Optional[int]:
    nums = []
    for p in root.iterdir():
        if not p.is_dir():
            continue
        m = re.match(r"^build\s*stage\s*(\d+)$", p.name, re.IGNORECASE)
        if m:
            nums.append(int(m.group(1)))
    return max(nums) if nums else None


def discover_tests(root: Path, include_tmp: bool = False) -> List[Path]:
    base = root / "uxm" / "tests"
    if not base.exists():
        return []
    seen = set()
    ordered: List[Path] = []
    for d in TEST_DIR_PRIORITY:
        full = root / d
        if full.exists():
            for p in sorted(full.glob("*.uxm"), key=lambda x: x.name.lower()):
                if not include_tmp and p.name.startswith("_"):
                    continue
                rp = p.resolve()
                if rp not in seen:
                    seen.add(rp)
                    ordered.append(p)
    for p in sorted(base.rglob("*.uxm"), key=lambda x: relpath(x, root).lower()):
        if not include_tmp and p.name.startswith("_"):
            continue
        rp = p.resolve()
        if rp not in seen:
            seen.add(rp)
            ordered.append(p)
    return ordered


def parse_smoke_expected(root: Path, stage: int) -> Tuple[Dict[str, str], List[Path]]:
    expected: Dict[str, str] = {}
    smoke_files: List[Path] = []
    exact = root / f"run_stage{stage}_smoke.bat"
    if exact.exists():
        smoke_files.append(exact)
    # Stage 12 için stage12 smoke yoksa en yeni stage smoke'u referans olarak kullan.
    all_smokes = []
    for p in root.glob("run_stage*_smoke.bat"):
        m = re.search(r"run_stage(\d+)_smoke\.bat", p.name, re.IGNORECASE)
        if m:
            all_smokes.append((int(m.group(1)), p))
    for st, p in sorted(all_smokes, reverse=True):
        if p not in smoke_files and st <= stage:
            smoke_files.append(p)
            break
    # Diğer smoke dosyaları varsa sonradan bilgi amaçlı listelenir, otomatik çalıştırılmaz.
    for sm in smoke_files:
        try:
            text = sm.read_text(encoding="utf-8", errors="replace")
        except Exception:
            text = sm.read_text(errors="replace")
        for line in text.splitlines():
            m = RUN_EXPECT_RE.match(line)
            if m:
                src = m.group(1).strip().strip('"')
                exp = m.group(3).strip().strip('"')
                expected[src.replace("/", "\\")] = exp
    # Harici expected CSV desteği.
    for csv_file in sorted(root.glob("expected_outputs*.csv")):
        try:
            with csv_file.open("r", encoding="utf-8-sig", errors="replace", newline="") as f:
                rows = list(csv.reader(f))
        except Exception:
            continue
        if not rows:
            continue
        header = [h.strip().lower() for h in rows[0]]
        has_header = any(h in header for h in ("test", "src", "source", "test_path", "expected"))
        data_rows = rows[1:] if has_header else rows
        if has_header:
            def idx(names, default):
                for n in names:
                    if n in header:
                        return header.index(n)
                return default
            i_test = idx(["test", "src", "source", "test_path", "path"], 0)
            i_exp = idx(["expected", "expect", "expected_output", "compact"], 1)
        else:
            i_test, i_exp = 0, 1
        for row in data_rows:
            if len(row) > max(i_test, i_exp) and row[i_test].strip():
                expected[row[i_test].strip().replace("/", "\\")] = row[i_exp].strip()
    return expected, smoke_files


def load_legacy_history(root: Path) -> Dict[str, List[float]]:
    hist = root / "test_history.csv"
    data: Dict[str, List[float]] = {}
    if not hist.exists():
        return data
    try:
        with hist.open("r", encoding="utf-8-sig", errors="replace", newline="") as f:
            for row in csv.reader(f):
                if len(row) < 3:
                    continue
                test = row[1].strip()
                try:
                    dur = float(str(row[2]).replace(",", "."))
                except Exception:
                    continue
                if test and dur >= 0:
                    data.setdefault(test.replace("/", "\\"), []).append(dur)
    except Exception:
        pass
    return data


def summarize_duration(values: List[float]) -> Dict[str, float]:
    if not values:
        return {"count": 0, "mean": 0.0, "std": 0.0, "min": 0.0, "max": 0.0, "last": 0.0}
    return {
        "count": len(values),
        "mean": statistics.mean(values),
        "std": statistics.stdev(values) if len(values) > 1 else 0.0,
        "min": min(values),
        "max": max(values),
        "last": values[-1],
    }


def perf_status(test_rel: str, duration: float, hist: Dict[str, List[float]]) -> Tuple[str, str]:
    vals = hist.get(test_rel.replace("/", "\\"), [])
    if not vals:
        return "NEW", "gecmis veri yok"
    s = summarize_duration(vals)
    mean = s["mean"]
    std = s["std"]
    hard_limit = max(s["max"] + max(0.25, 2.0 * std), mean * 1.35)
    warn_limit = max(mean * 1.20, s["max"] + 0.10)
    if duration > hard_limit:
        return "SLOW_CRITICAL", f"sure {duration:.3f} sn; ort {mean:.3f}, max {s['max']:.3f}, std {std:.3f}"
    if duration > warn_limit:
        return "SLOW_WARN", f"sure {duration:.3f} sn; ort {mean:.3f}, max {s['max']:.3f}"
    if duration < mean * 0.80:
        return "FAST", f"sure {duration:.3f} sn; ort {mean:.3f}"
    return "OK", f"sure {duration:.3f} sn; ort {mean:.3f}"


def update_stats_summary(root: Path, current_rows: List[Dict[str, str]]) -> None:
    hist = load_legacy_history(root)
    # current_rows henüz hist'e append edilmiş olabilir; direkt history'den oku daha sağlam.
    rows = []
    for test, vals in sorted(hist.items()):
        s = summarize_duration(vals)
        rows.append([
            test,
            s["count"],
            f"{s['mean']:.4f}",
            f"{s['last']:.4f}",
            f"{s['min']:.4f}",
            f"{s['max']:.4f}",
            f"{s['std']:.4f}",
        ])
    write_csv(root / "test_stats_summary.csv", ["Test_Adi", "Calistirma_Sayisi", "Ortalama_Sure", "Son_Sure", "Min_Sure", "Max_Sure", "Std_Sapma"], rows)


def try_write_xlsx_from_csv(csv_path: Path, xlsx_path: Path) -> Tuple[bool, str]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except Exception as exc:
        return False, f"openpyxl yok veya yuklenemedi: {exc!r}"
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "UXM_Test_Results"
        with csv_path.open("r", encoding="utf-8-sig", errors="replace", newline="") as f:
            for r_i, row in enumerate(csv.reader(f), 1):
                ws.append([xlsx_safe(col) for col in row])
                if r_i == 1:
                    for cell in ws[r_i]:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill("solid", fgColor="366092")
                        cell.alignment = Alignment(horizontal="center")
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            max_len = 0
            for cell in ws[letter]:
                max_len = max(max_len, len(str(cell.value or "")))
            ws.column_dimensions[letter].width = min(max(10, max_len + 2), 48)
        wb.save(xlsx_path)
        return True, "OK"
    except Exception as exc:
        err_path = xlsx_path.with_suffix(xlsx_path.suffix + ".xlsx_error.txt")
        write_text(err_path, "XLSX raporu yazilamadi; CSV ve raw log gecerlidir.\n" + repr(exc) + "\n")
        return False, repr(exc)


def run_existing_flow(args) -> int:
    root = Path(args.root).resolve()
    stage = int(args.stage) if args.stage is not None else ((detect_stage_from_build_names(root) or 11) + 1)
    result_no = next_sonuc_number(root)
    stamp = now_stamp()
    run_dir = root / "stage_runs" / f"stage_{stage}_{stamp}"
    logs_dir = run_dir / "logs"
    ensure_dir(logs_dir)
    sonuc_path = root / f"sonuc{result_no}.txt"
    run_sonuc_path = run_dir / f"sonuc{result_no}.txt"

    expected_map, smoke_files = parse_smoke_expected(root, stage)
    tests = discover_tests(root, include_tmp=args.include_tmp)
    old_hist = load_legacy_history(root)
    start_all = time.perf_counter()

    raw_lines = []
    raw_lines.append(f"UXM_EXISTING_FLOW_MANAGER stage={stage} sonuc={sonuc_path.name} start={now_human()}\n")
    raw_lines.append(f"ROOT={root}\n")
    raw_lines.append(f"TEST_COUNT={len(tests)}\n")
    raw_lines.append(f"EXPECTED_COUNT={len(expected_map)}\n")

    print(f"UXM mevcut akış runner başladı: stage={stage}, log={sonuc_path.name}, test={len(tests)}")

    smoke_status = "SKIPPED"
    smoke_elapsed = 0.0
    if not args.skip_smoke and smoke_files:
        sm = smoke_files[0]
        cmd = f'call "{sm.name}"'
        print(f"[SMOKE] {sm.name}")
        code, out, err, smoke_elapsed = run_shell(cmd, cwd=root, timeout=args.smoke_timeout)
        smoke_log = logs_dir / f"smoke_{sm.stem}.log"
        write_text(smoke_log, out + ("\n[STDERR]\n" + err if err else ""))
        raw_lines.append(f"SMOKE_START@{sm.name}@{now_human()}\n")
        raw_lines.append(out)
        if err:
            raw_lines.append("\n[SMOKE STDERR]\n" + err + "\n")
        raw_lines.append(f"SMOKE_END@{sm.name}@code={code}@elapsed={smoke_elapsed:.3f}\n")
        if code != 0:
            smoke_status = "FAIL"
            print(f"[SMOKE FAIL] {sm.name} ({smoke_elapsed:.2f} sn)")
            write_text(sonuc_path, "".join(raw_lines))
            shutil.copy2(sonuc_path, run_sonuc_path)
            return 1
        smoke_status = "OK"
        print(f"[SMOKE OK] {sm.name} ({smoke_elapsed:.2f} sn)")
    elif not smoke_files:
        raw_lines.append("SMOKE_SKIPPED@no_smoke_file\n")

    # Derleyiciyi derle.
    build_duration = 0.0
    build_code = 0
    if not args.skip_build:
        print("[BUILD] build_native.bat")
        build_code, build_out, build_err, build_duration = run_shell('call "build_native.bat"', cwd=root, timeout=args.build_timeout)
        write_text(logs_dir / "build_native.log", build_out + ("\n[STDERR]\n" + build_err if build_err else ""))
        raw_lines.append(f"START_BUILD@{now_human()}\n")
        raw_lines.append(build_out)
        if build_err:
            raw_lines.append("\n[BUILD STDERR]\n" + build_err + "\n")
        raw_lines.append(f"END_BUILD@code={build_code}@elapsed={build_duration:.3f}\n")
        if build_code != 0:
            print("[BUILD FAIL] build_native.bat")
            write_text(sonuc_path, "".join(raw_lines))
            shutil.copy2(sonuc_path, run_sonuc_path)
            return 1

    # Testleri çalıştır.
    result_header = [
        "timestamp", "stage", "result_no", "index", "total", "test_path", "status", "return_code",
        "duration_s", "build_duration_s", "expected", "expected_ok", "perf_status", "perf_note", "log_file",
    ]
    result_rows: List[List] = []
    current_rows: List[Dict[str, str]] = []
    failures = 0
    warnings = 0

    for idx, test in enumerate(tests, 1):
        test_rel = relpath(test, root)
        log_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", test_rel).replace(".uxm", ".log")
        test_log = logs_dir / log_name
        cmd = f'call "build_one_native.bat" "{test_rel}"'
        if args.test_exe_mode == "program":
            cmd += " -x"
        t0 = time.perf_counter()
        code, out, err, elapsed = run_shell(cmd, cwd=root, timeout=args.test_timeout)
        duration = time.perf_counter() - t0
        combined = out + ("\n[STDERR]\n" + err if err else "")
        write_text(test_log, combined)

        expect = expected_map.get(test_rel, expected_map.get(test_rel.replace("/", "\\"), ""))
        expected_ok = ""
        if expect:
            expected_ok = "OK" if expect in compact_text(combined) else "FAIL"
            if expected_ok == "FAIL":
                code = code if code != 0 else 90
        status = "BASARILI" if code == 0 else "BASARISIZ"
        if status != "BASARILI":
            failures += 1
        pstat, pnote = perf_status(test_rel, duration, old_hist)
        if pstat in ("SLOW_WARN", "SLOW_CRITICAL"):
            warnings += 1
        print(f"[{idx:03}/{len(tests):03}] {status} {test_rel} ({duration:.2f} sn) {pstat}")

        raw_lines.append("---------------------------------------\n")
        raw_lines.append(f"DATA_START@{test_rel}@{now_human()}\n")
        raw_lines.append(combined)
        raw_lines.append(f"\nRESULT@[{status}]@{test_rel}@code={code}@elapsed={duration:.3f}@expected={expect}@expected_ok={expected_ok}\n")
        raw_lines.append(f"DATA_END@{test_rel}@{now_human()}\n")

        row = [
            now_human(), stage, result_no, idx, len(tests), test_rel, status, code,
            f"{duration:.4f}", f"{build_duration:.4f}", expect, expected_ok, pstat, pnote,
            relpath(test_log, root),
        ]
        result_rows.append(row)
        current_rows.append({"test": test_rel, "duration": f"{duration:.4f}"})

        # Eski test_history.csv formatını bozma: tarih, test, süre, build_süre. Başlıksız.
        with (root / "test_history.csv").open("a", encoding="utf-8-sig", newline="") as f:
            csv.writer(f).writerow([now_human(), test_rel, f"{duration:.4f}", f"{build_duration:.4f}"])

    total_duration = time.perf_counter() - start_all
    raw_lines.append(f"END_SESSION@{now_human()}@elapsed={total_duration:.3f}@failures={failures}@warnings={warnings}\n")
    write_text(sonuc_path, "".join(raw_lines))
    shutil.copy2(sonuc_path, run_sonuc_path)

    results_csv = run_dir / "test_results.csv"
    write_csv(results_csv, result_header, result_rows)
    xlsx_ok, xlsx_msg = try_write_xlsx_from_csv(results_csv, run_dir / "test_results.xlsx")

    # Güncel istatistikleri yaz.
    update_stats_summary(root, current_rows)
    append_csv(
        root / "build_time_history.csv",
        ["timestamp", "stage", "result_no", "smoke_status", "smoke_s", "build_s", "full_test_s", "total_s", "test_count", "failures", "warnings"],
        [now_human(), stage, result_no, smoke_status, f"{smoke_elapsed:.4f}", f"{build_duration:.4f}", f"{total_duration - build_duration - smoke_elapsed:.4f}", f"{total_duration:.4f}", len(tests), failures, warnings],
    )

    summary = []
    summary.append(f"# UXM Stage {stage} Koşu Özeti\n")
    summary.append(f"- Tarih: {now_human()}\n")
    summary.append(f"- Root: `{root}`\n")
    summary.append(f"- Sonuç logu: `{sonuc_path.name}`\n")
    summary.append(f"- Smoke: {smoke_status} ({smoke_elapsed:.2f} sn)\n")
    summary.append(f"- Build süresi: {build_duration:.2f} sn\n")
    summary.append(f"- Test sayısı: {len(tests)}\n")
    summary.append(f"- Başarısız: {failures}\n")
    summary.append(f"- Süre uyarısı: {warnings}\n")
    summary.append(f"- Toplam süre: {total_duration:.2f} sn\n")
    summary.append(f"- XLSX raporu: {'OK' if xlsx_ok else 'YAZILAMADI'} — {xlsx_msg}\n")
    summary.append(f"- Test EXE modu: `{args.test_exe_mode}`\n")
    summary.append("\n## En yavaş 10 test\n\n")
    slow_sorted = sorted(result_rows, key=lambda r: float(r[8]), reverse=True)[:10]
    summary.append("| Test | Süre sn | Perf | Not |\n|---|---:|---|---|\n")
    for r in slow_sorted:
        summary.append(f"| `{r[5]}` | {r[8]} | {r[12]} | {r[13]} |\n")
    if failures:
        summary.append("\n## Başarısız testler\n\n")
        for r in result_rows:
            if r[6] != "BASARILI":
                summary.append(f"- `{r[5]}` code={r[7]} expected={r[10]} expected_ok={r[11]} log=`{r[14]}`\n")
    write_text(run_dir / "STAGE_RUN_SUMMARY.md", "".join(summary))

    # Stage build adı sadece raporlanır; taşıma toparla subcommand ile yapılır.
    build_stage_name = f"build stage {stage}"
    write_text(run_dir / "BUILD_STAGE_NAME.txt", build_stage_name + "\n")
    print(f"Bitti: failures={failures}, warnings={warnings}, summary={run_dir / 'STAGE_RUN_SUMMARY.md'}")
    print(f"Son build rapor adı: {build_stage_name}")
    return 1 if failures else 0


def audit_existing_files(args) -> int:
    root = Path(args.root).resolve()
    stage = int(args.stage) if args.stage is not None else 12
    out_dir = root / "stage_runs" / f"stage_{stage}_audit_{now_stamp()}"
    ensure_dir(out_dir)
    rows = []

    def add(kind, path, role, status, action, note=""):
        rows.append([kind, str(path), role, status, action, note])

    # BAT dosyaları.
    known_bat = {
        "build_native.bat": ("compiler build", "AKTIF", "koru", "fbc yolu var; yoksa PATH'teki fbc kullanılıyor"),
        "build_one_native.bat": ("tek .uxm derle+link+çalıştır", "AKTIF", "koru; pause patch gerekirse küçük yama", "optimizer için named exe modunda -x kullanma"),
        "run_stage11_smoke.bat": ("smoke + expected-output kapısı", "AKTIF", "önce çalıştır", "Stage 12 smoke yoksa en yeni smoke olarak kullanılır"),
        "run_tests_native.bat": ("basit full test", "YEDEK", "aktif hatta zorunlu değil", "istatistik üretmez"),
        "rtx.bat": ("eski sonucN logger", "LEGACY", "toparlayıcı ile emekli edilebilir", "@ işaretli veri üretir"),
        "rtxz.bat": ("eski sonucN logger varyantı", "LEGACY", "toparlayıcı ile emekli edilebilir", "@ işaretli veri üretir"),
        "runalltests.bat": ("eski full test", "HATALI/LEGACY", "emekli et veya parantez düzelt", "if/for kapanış parantezi eksik görünüyor"),
        "run_opt.bat": ("optimizer zinciri", "DÜZELTME GEREK", "manager opt sırasını kullan", "uxm_optimizer_pro.py çağırıyor ama zipte uxm_optimizer_pro2.py var"),
    }
    for p in sorted(root.glob("*.bat")):
        role, status, action, note = known_bat.get(p.name, ("bat/script", "BILINMIYOR", "incele", ""))
        add("bat", p.name, role, status, action, note)

    known_py_notes = {
        "zekiassop.py": ("ASM örüntü analizi + optimize ASM üretimi", "AKTIF AMA HARD-CODED", "manager import ederek çalıştır", "__main__ eski C:\\Users\\mete yoluna sabit"),
        "UXM_Heavy_Asm_Optimizer.py": ("optimizer öneri raporu", "ZAYIF", "kural kitabı genişlet", "rules=[] boş; strateji_kitabi_v2 Kural Sayısı 0 üretiyor"),
        "build_optimized.py": ("yeni_optimize_asm derleme/link/çalıştırma", "AKTIF", "opt zincirinde çalıştır", "FBC yolu sabit ama mevcutsa sorun yok"),
        "uxm_optimizer_pro2.py": ("orijinal/opt exe kıyas + sqlite", "AKTIF", "run_opt son adımı bu olmalı", "run_opt yanlışlıkla pro.py çağırıyor"),
        "stat.py": ("eski log parser", "LEGACY", "emekli et", "sonuc.txt arıyor; yeni sonucN akışıyla uyumsuz"),
        "sts.py": ("eski log parser", "LEGACY", "emekli et veya sadece referans", "eski format varsayımları var"),
        "stsx.py": ("eski log parser/rapor", "LEGACY", "emekli et veya sadece referans", "yeni manager doğrudan CSV üretir"),
        "UXMPerformansAnalizatoru.py": ("test_history Excel raporu", "AKTIF/YEDEK", "koru", "pandas/openpyxl ister"),
        "asmoptimizer.py": ("isim yanıltıcı performans analiz kopyası", "DUPLICATE", "emekli adayı", "ASM optimizer değil; UXMPerformansAnalizatoru ile aynı iş"),
        "uxm_analizor.py": ("servis/analiz", "ANALIZ", "koru", "rapor üretmek için kullanılabilir"),
        "uxm_analizor2.py": ("servis/analiz", "ANALIZ", "koru", "rapor üretmek için kullanılabilir"),
        "uxm_analizor(birlesik).py": ("birleşik analiz", "ANALIZ", "koru", "daha dolu analiz dosyası"),
    }
    for p in sorted(root.glob("*.py")):
        role, status, action, note = known_py_notes.get(p.name, ("python script", "BILINMIYOR", "incele", ""))
        add("py", p.name, role, status, action, note)

    for p in sorted(root.glob("*.csv")):
        add("csv", p.name, "geçmiş/veri", "KORU", "aktif hatta veri kaynağı", "test_history ve stats summary korunmalı")
    for p in sorted(root.glob("*.xlsx")):
        add("xlsx", p.name, "önceki rapor çıktısı", "ARSIV", "toparlayıcı ile rapor arşivine taşı", "aktif girdi değil")
    if (root / "optimizasyon" / "strateji_kitabi_v2.txt").exists():
        text = (root / "optimizasyon" / "strateji_kitabi_v2.txt").read_text(encoding="utf-8", errors="replace")
        if "Kural Sayısı: 0" in text:
            add("optimizer", "optimizasyon/strateji_kitabi_v2.txt", "ağır optimizer kural kitabı", "ZAYIF", "kural seti doldur", "mevcut dosyada Kural Sayısı: 0")

    write_csv(out_dir / "existing_file_audit.csv", ["kind", "path", "role", "status", "action", "note"], rows)
    md = [f"# UXM V33 Mevcut Dosya Akış Denetimi — Stage {stage}\n\n"]
    md.append("Bu rapor yeni bat dosyası üretmeden, mevcut dosyaların görev sırasını ve durumunu gösterir.\n\n")
    md.append("| Tür | Dosya | Durum | Görev | Öneri | Not |\n|---|---|---|---|---|---|\n")
    for kind, path, role, status, action, note in rows:
        md.append(f"| {kind} | `{path}` | {status} | {role} | {action} | {note} |\n")
    write_text(out_dir / "MEVCUT_DOSYA_AKIS_RAPORU.md", "".join(md))
    print(f"Audit yazıldı: {out_dir}")
    return 0


RULEBOOK = [
    {"id": "P001", "tier": "safe", "detect": "mov r64, 0", "replace": "xor r64, r64", "guard": "sadece aynı register ve immediate 0; flag etkisi kabul edilebilir yerde", "note": "rax/rbx/rcx/rdx/r8-r15 için"},
    {"id": "P002", "tier": "safe", "detect": "add/sub reg, 1", "replace": "inc/dec reg", "guard": "CF bayrağı kullanılmıyorsa", "note": "UXM flag hesaplaması ayrı yapılıyorsa güvenli"},
    {"id": "P003", "tier": "safe", "detect": "jmp next_label", "replace": "sil", "guard": "hedef bir sonraki satır etiketi ise", "note": "doğrudan düşüş"},
    {"id": "P004", "tier": "guarded", "detect": "push rax ... pop rax", "replace": "r8-r11 shadow", "guard": "arada rax değişiyorsa push/pop silinmez; sadece shadow yapılır", "note": "mevcut zekiassop burada fazla agresif"},
    {"id": "P005", "tier": "emitter", "detect": "tekrar eden ux_flags load/store", "replace": "flag shadow register", "guard": "basic block giriş/çıkışında sync", "note": "runtime semantics bozulmamalı"},
    {"id": "P006", "tier": "emitter", "detect": "cmp r10, TAPE_CELLS her erişimde", "replace": "loop precheck/hoist", "guard": "sabit count ve monoton indeks varsa", "note": "tensor/matrix döngülerinde güçlü kazanç"},
    {"id": "P007", "tier": "safe", "detect": "mov reg, reg", "replace": "sil", "guard": "aynı register", "note": "gereksiz self move"},
    {"id": "P008", "tier": "safe", "detect": "add reg, 0 / sub reg, 0", "replace": "sil", "guard": "flag etkisi kullanılmıyorsa", "note": "bazı asm flag yan etkisine dikkat"},
    {"id": "P009", "tier": "safe", "detect": "imul reg, reg, 2/4/8/16/32", "replace": "lea veya shl", "guard": "signed overflow flag kullanılmıyorsa", "note": "adres hesaplarında uygun"},
    {"id": "P010", "tier": "guarded", "detect": "lea r11,[r12+idx] + mov byte [r11]", "replace": "mov byte [r12+idx]", "guard": "NASM addressing legal ve r11 sonradan kullanılmıyorsa", "note": "register basıncını düşürür"},
    {"id": "P011", "tier": "emitter", "detect": "CALL meta küçük sabit id", "replace": "inline service fast path", "guard": "servis yan etkileri tam biliniyorsa", "note": "@8/@16/@32 gibi sıcak servislerde"},
    {"id": "P012", "tier": "analysis", "detect": "jump density yüksek", "replace": "block layout / fallthrough reorder", "guard": "etiket hedefleri korunur", "note": "rapor önerisi; otomatik uygulama yok"},
]


def optimizer_flow(args) -> int:
    root = Path(args.root).resolve()
    stage = int(args.stage) if args.stage is not None else 12
    out_dir = root / "stage_runs" / f"stage_{stage}_optimizer_{now_stamp()}"
    ensure_dir(out_dir)
    opt_dir = root / "optimizasyon"
    ensure_dir(opt_dir)

    # Genişletilmiş kural kitabı raporu.
    write_text(opt_dir / "asm_optimizer_rulebook_expanded.json", json.dumps(RULEBOOK, ensure_ascii=False, indent=2))
    md = [f"# UXM ASM Optimizer Genişletilmiş Kural Kitabı — Stage {stage}\n\n"]
    md.append("Bu dosya mevcut optimizer dosyalarının yerine geçmez; zayıf kalan kural setini görünür yapar. Otomatik uygulama için `safe`, `guarded`, `emitter` ayrımı korunmalıdır.\n\n")
    md.append("| ID | Katman | Tespit | Dönüşüm | Koruma | Not |\n|---|---|---|---|---|---|\n")
    for r in RULEBOOK:
        md.append(f"| {r['id']} | {r['tier']} | `{r['detect']}` | `{r['replace']}` | {r['guard']} | {r['note']} |\n")
    write_text(opt_dir / "asm_optimizer_rulebook_expanded.md", "".join(md))

    # run_opt.bat denetimi.
    run_opt = root / "run_opt.bat"
    if run_opt.exists():
        text = run_opt.read_text(encoding="utf-8", errors="replace")
        if "uxm_optimizer_pro.py" in text and not (root / "uxm_optimizer_pro.py").exists():
            write_text(out_dir / "run_opt_bat_issue.txt", "run_opt.bat uxm_optimizer_pro.py çağırıyor; zip içinde uxm_optimizer_pro2.py var. Manager opt akışı pro2'yi kullanır.\n")

    commands = []
    # Hard-coded __main__ yolunu aşmak için modül içindeki sınıflar import edilerek çağrılır.
    if (root / "zekiassop.py").exists():
        commands.append(("zekiassop_import", 'python -c "import os, zekiassop; zekiassop.UXM_ASM_Intelligence(os.getcwd()).run()"'))
    if (root / "UXM_Heavy_Asm_Optimizer.py").exists():
        commands.append(("heavy_optimizer_import", 'python -c "import os, UXM_Heavy_Asm_Optimizer as h; print(h.UXM_Final_Optimizer(os.getcwd()).analyze_and_suggest()); print(h.UXM_Heavy_Optimizer(os.getcwd()).build_report())"'))
    if not args.analyze_only and (root / "build_optimized.py").exists():
        commands.append(("build_optimized", 'python build_optimized.py'))
    if not args.analyze_only and (root / "uxm_optimizer_pro2.py").exists():
        commands.append(("uxm_optimizer_pro2", 'python uxm_optimizer_pro2.py'))

    summary_rows = []
    for name, cmd in commands:
        print(f"[OPT] {name}")
        code, out, err, elapsed = run_shell(cmd, cwd=root, timeout=args.opt_timeout)
        log_path = out_dir / f"{name}.log"
        write_text(log_path, out + ("\n[STDERR]\n" + err if err else ""))
        summary_rows.append([name, cmd, code, f"{elapsed:.4f}", relpath(log_path, root)])
        if code != 0 and not args.continue_on_error:
            break

    # asm_intel_report içinden örüntü sayımı.
    intel = opt_dir / "asm_intel_report.txt"
    audit_rows = []
    if intel.exists():
        text = intel.read_text(encoding="utf-8", errors="replace")
        for r in RULEBOOK:
            needle = r["detect"].split()[0]
            audit_rows.append([r["id"], r["tier"], r["detect"], text.lower().count(needle.lower()), r["note"]])
    write_csv(out_dir / "optimizer_flow_summary.csv", ["step", "command", "code", "elapsed_s", "log"], summary_rows)
    write_csv(out_dir / "asm_rule_audit.csv", ["rule_id", "tier", "detect", "rough_hit_count", "note"], audit_rows)
    print(f"Optimizer akış özeti: {out_dir}")
    return 0 if all(str(r[2]) == "0" for r in summary_rows) else 1


def collect_moves_for_toparla(root: Path, stage: int, include_scripts: bool, move_build: bool) -> List[Tuple[Path, Path, str]]:
    base = root / "_UXM_EMEKLI"
    stamp = now_stamp()
    moves: List[Tuple[Path, Path, str]] = []

    # Eski raporlar ve sonuc logları.
    for p in root.glob("*.xlsx"):
        moves.append((p, base / "reports" / "xlsx" / stamp / p.name, "xlsx rapor arşivi"))
    for p in root.glob("sonuc*.txt"):
        moves.append((p, base / "logs" / "sonuc" / stamp / p.name, "sonuc log arşivi"))
    for p in root.glob("Performans_Raporu_*.xlsx"):
        moves.append((p, base / "reports" / "xlsx" / stamp / p.name, "performans rapor arşivi"))

    # Geçici dosyalar.
    for p in root.rglob("_tmp*"):
        if p.is_file():
            moves.append((p, base / "tmp" / stamp / relpath(p, root), "geçici dosya"))

    # Build çıktısı.
    if move_build and (root / "build").exists():
        moves.append((root / "build", base / "builds" / f"build_stage_{stage}_{stamp}", "işi biten build klasörü"))
    for p in root.iterdir():
        if p.is_dir() and re.match(r"^build\s*stage\s*\d+", p.name, re.IGNORECASE):
            moves.append((p, base / "builds" / stamp / p.name, "elle numaralanmış eski build"))

    # Scriptler sadece istenirse taşınır; default dokunma.
    if include_scripts:
        for name in ["runalltests.bat", "rtx.bat", "rtxz.bat", "stat.py", "sts.py", "stsx.py", "asmoptimizer.py"]:
            p = root / name
            if p.exists():
                moves.append((p, base / "legacy_scripts" / stamp / name, "legacy/duplicate script"))
    return moves


def toparla(args) -> int:
    root = Path(args.root).resolve()
    stage = int(args.stage) if args.stage is not None else 12
    moves = collect_moves_for_toparla(root, stage, include_scripts=args.include_scripts, move_build=args.move_build)
    out_dir = root / "_UXM_EMEKLI" / "manifests"
    ensure_dir(out_dir)
    stamp = now_stamp()
    manifest_csv = out_dir / f"toparlama_manifest_stage_{stage}_{stamp}.csv"
    rows = []
    for src, dst, reason in moves:
        rows.append([relpath(src, root), relpath(dst, root), reason, "APPLY" if args.apply else "DRY_RUN"])
    write_csv(manifest_csv, ["source", "destination", "reason", "mode"], rows)
    md = [f"# UXM Toparlama Manifesti — Stage {stage}\n\n"]
    md.append(f"Mod: {'APPLY' if args.apply else 'DRY-RUN'}\n\n")
    md.append("| Kaynak | Hedef | Sebep |\n|---|---|---|\n")
    for src, dst, reason in moves:
        md.append(f"| `{relpath(src, root)}` | `{relpath(dst, root)}` | {reason} |\n")
    write_text(out_dir / f"toparlama_manifest_stage_{stage}_{stamp}.md", "".join(md))

    if args.apply:
        for src, dst, reason in moves:
            if not src.exists():
                continue
            ensure_dir(dst.parent)
            if dst.exists():
                dst = dst.with_name(dst.stem + "_" + stamp + dst.suffix)
            shutil.move(str(src), str(dst))
        if args.move_build:
            ensure_dir(root / "build" / "asm")
            ensure_dir(root / "build" / "obj")
            ensure_dir(root / "build" / "exe")
    print(f"Toparlama manifesti: {manifest_csv}")
    if not args.apply:
        print("DRY-RUN: Dosyalar taşınmadı. Gerçek taşıma için --apply kullan.")
    return 0


def replace_once(text: str, old: str, new: str) -> Tuple[str, bool]:
    if old not in text:
        return text, False
    return text.replace(old, new, 1), True


def pause_patch(args) -> int:
    root = Path(args.root).resolve()
    targets = {
        "compiler": root / "uxm" / "core" / "compiler" / "native" / "uxm31_compiler_fb.bas",
        "main": root / "uxm" / "core" / "compiler" / "native" / "native_main.bas",
        "cli": root / "uxm" / "core" / "compiler" / "native" / "native_cli.bas",
        "emit": root / "uxm" / "core" / "compiler" / "native" / "native_asm_emit.bas",
        "runtime": root / "uxm" / "core" / "runtime" / "uxm31_runtime_fb_full.bas",
        "bat": root / "build_one_native.bat",
    }
    missing = [str(p) for p in targets.values() if not p.exists()]
    if missing:
        print("Eksik dosya var; patch uygulanmadı:")
        for m in missing:
            print(" -", m)
        return 1

    backup_dir = root / "_UXM_PATCH_BACKUP" / now_stamp()
    changes = []

    def patch_file(key: str, patch_func):
        p = targets[key]
        text = p.read_text(encoding="utf-8", errors="replace")
        new_text, changed, note = patch_func(text)
        changes.append([relpath(p, root), "CHANGED" if changed else "SKIP", note])
        if args.apply and changed:
            dst = backup_dir / relpath(p, root)
            ensure_dir(dst.parent)
            shutil.copy2(p, dst)
            p.write_text(new_text, encoding="utf-8")

    def patch_compiler(text: str):
        if "UXM_PauseAtEnd" in text:
            return text, False, "UXM_PauseAtEnd zaten var"
        old = "Dim Shared PragmaArgeWatch As Long\nDim Shared OutFF As Long"
        new = "Dim Shared PragmaArgeWatch As Long\nDim Shared UXM_PauseAtEnd As Long\nDim Shared OutFF As Long"
        nt, ok = replace_once(text, old, new)
        return nt, ok, "shared pause bayrağı eklendi" if ok else "hedef nokta bulunamadı"

    def patch_cli(text: str):
        changed = False
        note = []
        if "UXM_PauseAtEnd=0" not in text:
            old = "PragmaArgeWatch=0\n    ApplyMemoryModel()"
            new = "PragmaArgeWatch=0\n    UXM_PauseAtEnd=0\n    ApplyMemoryModel()"
            text, ok = replace_once(text, old, new)
            changed = changed or ok
            note.append("InitDefaults pause=0" if ok else "InitDefaults hedefi yok")
        if "#nopause" not in text and "#pause" not in text:
            old = "ElseIf InStr(low,\"#arge\")=1 Then\n                If InStr(low,\"json\")>0 Then PragmaArgeJson=1"
            new = "ElseIf InStr(low,\"#pause\")=1 Then\n                UXM_PauseAtEnd=1\n            ElseIf InStr(low,\"#nopause\")=1 Then\n                UXM_PauseAtEnd=0\n            ElseIf InStr(low,\"#arge\")=1 Then\n                If InStr(low,\"json\")>0 Then PragmaArgeJson=1"
            text, ok = replace_once(text, old, new)
            changed = changed or ok
            note.append("#pause/#nopause pragma" if ok else "pragma hedefi yok")
        return text, changed, "; ".join(note)

    def patch_main(text: str):
        if "UXM_PauseAtEnd" in text and "--pause" in text:
            return text, False, "CLI pause parse zaten var"
        old = "If Command(2)<>\"\" Then\n        OutAsm=TrimAll(Command(2))\n    Else\n        OutAsm=InFile+\".asm\"\n    End If"
        new = "If Command(2)<>\"\" Then\n        OutAsm=TrimAll(Command(2))\n    Else\n        OutAsm=InFile+\".asm\"\n    End If\n    Dim __uxArgI As Long\n    Dim __uxArg As String\n    __uxArgI=3\n    Do\n        __uxArg=TrimAll(Command(__uxArgI))\n        If __uxArg=\"\" Then Exit Do\n        If LCase(__uxArg)=\"--pause\" Or LCase(__uxArg)=\"-p\" Then UXM_PauseAtEnd=1\n        If LCase(__uxArg)=\"--no-pause\" Then UXM_PauseAtEnd=0\n        __uxArgI=__uxArgI+1\n    Loop"
        nt, ok = replace_once(text, old, new)
        return nt, ok, "Command(3+) --pause parse eklendi" if ok else "native_main hedefi bulunamadı"

    def patch_emit(text: str):
        changed = False
        note = []
        if "extern ux_pause_at_end" not in text:
            old = "EmitLine(\"extern ux_runtime_error\")"
            new = "EmitLine(\"extern ux_runtime_error\")\n    If UXM_PauseAtEnd<>0 Then EmitLine(\"extern ux_pause_at_end\")"
            text, ok = replace_once(text, old, new)
            changed = changed or ok
            note.append("extern eklendi" if ok else "extern hedefi yok")
        if "call ux_pause_at_end" not in text:
            old = "Sub EmitFooter()\n    EmitLine(\"__ux_ok_exit:\")\n    EmitLine(\"    add rsp, 40\")"
            new = "Sub EmitFooter()\n    EmitLine(\"__ux_ok_exit:\")\n    If UXM_PauseAtEnd<>0 Then EmitLine(\"    call ux_pause_at_end\")\n    EmitLine(\"    add rsp, 40\")"
            text, ok = replace_once(text, old, new)
            changed = changed or ok
            note.append("exit öncesi call eklendi" if ok else "footer hedefi yok")
        return text, changed, "; ".join(note)

    def patch_runtime(text: str):
        if "Sub ux_pause_at_end" in text:
            return text, False, "runtime pause fonksiyonu zaten var"
        insert = "Declare Sub ux_runtime_error(ByVal code As ULongInt)\n"
        repl = insert + "Declare Sub ux_pause_at_end()\n"
        text2, ok1 = replace_once(text, insert, repl)
        func = "\nSub ux_pause_at_end() Export\n    Print\n    Print \"[UXM] Program bitti. Devam etmek icin bir tusa basin...\"\n    Sleep\nEnd Sub\n"
        # Include'lardan önce ekle.
        marker = "#Include Once \"runtime_memory.bas\""
        if marker in text2:
            text2 = text2.replace(marker, func + "\n" + marker, 1)
            return text2, True, "runtime declare + exported pause sub eklendi"
        return text2, ok1, "declare eklendi ama include marker bulunamadı"

    def patch_bat(text: str):
        if "EXTRA_ARGS" in text and "%EXTRA_ARGS%" in text:
            return text, False, "build_one_native.bat ekstra arg zaten geçiriyor"
        old = 'build\\exe\\uxm_native.exe "%~1" "build\\asm\\%NAME%.asm"'
        new = 'set EXTRA_ARGS=\nif not "%~3"=="" set EXTRA_ARGS=%~3\nif not "%~4"=="" set EXTRA_ARGS=%EXTRA_ARGS% %~4\nif not "%~5"=="" set EXTRA_ARGS=%EXTRA_ARGS% %~5\nbuild\\exe\\uxm_native.exe "%~1" "build\\asm\\%NAME%.asm" %EXTRA_ARGS%'
        nt, ok = replace_once(text, old, new)
        return nt, ok, "build_one_native üçüncü argümandan itibaren compiler'a geçirir" if ok else "bat hedef satırı bulunamadı"

    patch_file("compiler", patch_compiler)
    patch_file("cli", patch_cli)
    patch_file("main", patch_main)
    patch_file("emit", patch_emit)
    patch_file("runtime", patch_runtime)
    patch_file("bat", patch_bat)

    report_dir = root / "_UXM_PATCH_BACKUP" / "manifests"
    ensure_dir(report_dir)
    write_csv(report_dir / f"pause_patch_manifest_{now_stamp()}.csv", ["file", "status", "note"], changes)
    for row in changes:
        print(row[1], row[0], "-", row[2])
    if args.apply:
        print(f"Patch uygulandı. Yedek: {backup_dir}")
        print("Sonra derle: call build_native.bat")
        print('Manuel EXE görmek için: call build_one_native.bat "uxm\\tests\\native\\test05_meta_add.uxm" -x --pause')
    else:
        print("DRY-RUN: Patch uygulanmadı. Gerçek uygulama için --apply kullan.")
    return 0


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="UXM V33 mevcut dosyaları sıraya koyan stage/optimizer/toparlama yöneticisi")
    parser.add_argument("command", choices=["run", "audit", "opt", "toparla", "pause-patch"], help="işlem türü")
    parser.add_argument("--root", default=".", help="UXMv33 kök klasörü")
    parser.add_argument("--stage", type=int, default=None, help="stage numarası")

    parser.add_argument("--skip-smoke", action="store_true")
    parser.add_argument("--skip-build", action="store_true")
    parser.add_argument("--include-tmp", action="store_true")
    parser.add_argument("--test-exe-mode", choices=["named", "program"], default="named", help="named: her teste ayrı exe; program: -x ile program.exe")
    parser.add_argument("--build-timeout", type=int, default=120)
    parser.add_argument("--smoke-timeout", type=int, default=300)
    parser.add_argument("--test-timeout", type=int, default=120)

    parser.add_argument("--analyze-only", action="store_true", help="opt: optimize asm derleme/kıyas adımlarını atla")
    parser.add_argument("--continue-on-error", action="store_true", help="opt: hata olsa da sonraki adıma geç")
    parser.add_argument("--opt-timeout", type=int, default=1800)

    parser.add_argument("--dry-run", action="store_true", help="toparla/pause-patch için sadece plan")
    parser.add_argument("--apply", action="store_true", help="toparla/pause-patch gerçek uygula")
    parser.add_argument("--include-scripts", action="store_true", help="toparla: legacy scriptleri de taşı")
    parser.add_argument("--move-build", action="store_true", help="toparla: build klasörünü emekliye taşı ve boş build oluştur")

    args = parser.parse_args(argv)
    if args.command in ("toparla", "pause-patch") and not args.apply:
        args.dry_run = True

    if args.command == "run":
        return run_existing_flow(args)
    if args.command == "audit":
        return audit_existing_files(args)
    if args.command == "opt":
        return optimizer_flow(args)
    if args.command == "toparla":
        return toparla(args)
    if args.command == "pause-patch":
        return pause_patch(args)
    return 2


if __name__ == "__main__":
    raise SystemExit(main())
